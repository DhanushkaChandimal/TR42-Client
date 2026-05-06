"""Excel export service.

Builds polished .xlsx workbooks for two pages:
  - Analytics: KPIs, vendor performance, invoice pipeline, outstanding,
    cost by service, WO status, MSA expiries
  - Invoices: every invoice with vendor + WO context, line items on a
    second sheet, summary on a third

Shared formatting helpers up top so both workbooks look consistent:
bold colored headers, frozen first row, auto-filter, currency / date
formatting, sized columns.
"""
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from statistics import quantiles, median

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.models.client import Client
from app.models.invoice import Invoice
from app.models.line_item import LineItem
from app.models.msa import Msa
from app.models.ticket import Ticket
from app.models.vendor import Vendor
from app.models.workorder import WorkOrder


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="EEF2FF")
SUBHEADER_FONT = Font(bold=True, color="1F3A8A", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F2937")
META_FONT = Font(italic=True, color="6B7280", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
SEVERITY_FILLS = {
    "REJECTED": PatternFill("solid", fgColor="FEE2E2"),
    "PAID": PatternFill("solid", fgColor="DCFCE7"),
    "APPROVED": PatternFill("solid", fgColor="DCFCE7"),
    "PENDING": PatternFill("solid", fgColor="FEF3C7"),
    "DRAFT": PatternFill("solid", fgColor="F3F4F6"),
    "SUBMITTED": PatternFill("solid", fgColor="DBEAFE"),
}


def _write_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = META_FONT
    ws["A3"] = (
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A3"].font = META_FONT


def _write_header(ws, row, columns):
    for col_idx, (label, _width, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def _apply_widths(ws, columns, start_col=1):
    for col_idx, (_label, width, _fmt) in enumerate(columns, start=start_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_row(ws, row_idx, columns, values, status_col=None):
    fill = None
    if status_col is not None:
        status_value = values[status_col]
        if status_value in SEVERITY_FILLS:
            fill = SEVERITY_FILLS[status_value]
    for col_idx, ((_, _, fmt), value) in enumerate(zip(columns, values), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if fmt:
            cell.number_format = fmt
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _finalize_data_sheet(ws, header_row, columns, data_row_count):
    if data_row_count == 0:
        return
    last_col = get_column_letter(len(columns))
    last_row = header_row + data_row_count
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    ws.freeze_panes = ws[f"A{header_row + 1}"]


def _naive(value):
    """Strip tzinfo from datetime/date so openpyxl can serialize it.

    Postgres timestamptz columns come back as aware datetimes; Excel cells
    only accept naive datetimes.
    """
    if value is None:
        return None
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _to_workbook_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _vendor_name(v):
    if not v:
        return "Unknown"
    return v.company_name or getattr(v, "name", None) or "Unknown"


# ---------------------------------------------------------------------------
# Analytics workbook
# ---------------------------------------------------------------------------


def build_analytics_workbook(start=None, end=None):
    wb = Workbook()
    wb.remove(wb.active)

    tickets = [
        t for t in db.session.execute(select(Ticket)).scalars().all()
        if not (start or end) or _in_range(t.created_at, start, end)
    ]
    vendors = db.session.execute(select(Vendor)).scalars().all()
    invoices = [
        i for i in db.session.execute(select(Invoice)).scalars().all()
        if not (start or end) or _in_range(i.invoice_date, start, end)
    ]
    work_orders = [
        w for w in db.session.execute(select(WorkOrder)).scalars().all()
        if not (start or end) or _in_range(w.created_at, start, end)
    ]
    msas = db.session.execute(select(Msa)).scalars().all()

    vendors_by_id = {v.id: v for v in vendors}
    workorders_by_id = {w.id: w for w in work_orders}

    _write_summary_sheet(wb, tickets, vendors, invoices, work_orders, msas, period=_period_label(start, end))
    _write_vendor_performance_sheet(wb, tickets, vendors_by_id)
    _write_invoice_pipeline_sheet(wb, invoices)
    _write_outstanding_sheet(wb, invoices, vendors_by_id)
    _write_cost_by_service_sheet(wb, invoices, workorders_by_id)
    _write_wo_status_sheet(wb, work_orders)
    _write_msa_expiry_sheet(wb, msas, vendors_by_id, work_orders)

    return _to_workbook_bytes(wb)


def _write_summary_sheet(ws_wb, tickets, vendors, invoices, work_orders, msas, period=None):
    ws = ws_wb.create_sheet("Overview")
    subtitle = "Top-line KPIs across tickets, vendors, invoices, work orders."
    if period:
        subtitle = f"{subtitle}  ·  Period: {period}"
    _write_title(
        ws,
        "Analytics Overview",
        subtitle,
    )

    approved = sum(1 for t in tickets if _enum_value(t.status) == "APPROVED")
    rejected = sum(1 for t in tickets if _enum_value(t.status) == "REJECTED")
    reviewed = approved + rejected
    approval_rate = (approved / reviewed) if reviewed else 0
    active_wo = sum(
        1
        for w in work_orders
        if _enum_value(getattr(w, "current_status", None)) not in ("CANCELLED", "CLOSED")
    )

    rows = [
        ("Total tickets", len(tickets), None),
        ("Approved tickets", approved, None),
        ("Rejected tickets", rejected, None),
        ("Approval rate", approval_rate, "0.0%"),
        ("Vendors", len(vendors), None),
        ("Active work orders", active_wo, None),
        ("Total invoices", len(invoices), None),
        (
            "Total invoiced ($)",
            sum(float(i.total_amount or 0) for i in invoices),
            "$#,##0.00",
        ),
        ("MSAs on file", len(msas), None),
    ]

    header_row = 5
    cols = [("Metric", 32, None), ("Value", 22, None)]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    for i, (label, value, fmt) in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=label).border = THIN_BORDER
        cell = ws.cell(row=i, column=2, value=value)
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_vendor_performance_sheet(ws_wb, tickets, vendors_by_id):
    ws = ws_wb.create_sheet("Vendor Performance")
    _write_title(ws, "Vendor Performance", "Ticket counts, approval rates, and average approval time per vendor.")

    by_vendor = defaultdict(list)
    for t in tickets:
        if t.vendor_id:
            by_vendor[t.vendor_id].append(t)

    rows = []
    for vendor_id, ts in by_vendor.items():
        v = vendors_by_id.get(vendor_id)
        approved = [t for t in ts if _enum_value(t.status) == "APPROVED"]
        rejected = [t for t in ts if _enum_value(t.status) == "REJECTED"]
        pending = [t for t in ts if _enum_value(t.status) == "PENDING_APPROVAL"]
        reviewed = len(approved) + len(rejected)
        rejection_rate = len(rejected) / reviewed if reviewed else 0

        approval_hours_total = 0.0
        approval_count = 0
        for t in approved:
            if t.created_at and t.approved_at:
                delta = (t.approved_at - t.created_at).total_seconds() / 3600
                approval_hours_total += delta
                approval_count += 1
        avg_hours = approval_hours_total / approval_count if approval_count else None

        rows.append(
            [
                v.company_name if v and v.company_name else (v.name if v else vendor_id[:8]),
                len(ts),
                len(approved),
                len(rejected),
                len(pending),
                rejection_rate,
                avg_hours,
            ]
        )

    rows.sort(key=lambda r: (r[5], r[3]), reverse=True)

    header_row = 5
    cols = [
        ("Vendor", 30, None),
        ("Total tickets", 14, "#,##0"),
        ("Approved", 12, "#,##0"),
        ("Rejected", 12, "#,##0"),
        ("Pending", 12, "#,##0"),
        ("Rejection rate", 16, "0.0%"),
        ("Avg approval (hrs)", 18, "0.0"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    for i, row in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, cols, row)
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_invoice_pipeline_sheet(ws_wb, invoices):
    ws = ws_wb.create_sheet("Invoice Pipeline")
    _write_title(ws, "Invoice Pipeline", "Invoice counts and dollar totals by status.")

    buckets = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for inv in invoices:
        status = _enum_value(inv.invoice_status) or "UNKNOWN"
        buckets[status]["count"] += 1
        buckets[status]["amount"] += float(inv.total_amount or 0)

    header_row = 5
    cols = [
        ("Status", 18, None),
        ("Count", 12, "#,##0"),
        ("Total ($)", 18, "$#,##0.00"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    rows = sorted(buckets.items(), key=lambda kv: kv[1]["amount"], reverse=True)
    for i, (status, slot) in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, cols, [status, slot["count"], slot["amount"]], status_col=0)
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_outstanding_sheet(ws_wb, invoices, vendors_by_id):
    ws = ws_wb.create_sheet("Outstanding by Vendor")
    _write_title(ws, "Outstanding Invoices by Vendor", "Anything not PAID and not REJECTED counts as outstanding.")

    by_vendor = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for inv in invoices:
        status = _enum_value(inv.invoice_status)
        if status in ("PAID", "REJECTED"):
            continue
        vid = inv.vendor_id or "unknown"
        by_vendor[vid]["count"] += 1
        by_vendor[vid]["amount"] += float(inv.total_amount or 0)

    header_row = 5
    cols = [
        ("Vendor", 30, None),
        ("Open invoices", 14, "#,##0"),
        ("Outstanding ($)", 20, "$#,##0.00"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    rows = sorted(by_vendor.items(), key=lambda kv: kv[1]["amount"], reverse=True)
    for i, (vid, slot) in enumerate(rows, start=header_row + 1):
        v = vendors_by_id.get(vid)
        name = (
            v.company_name if v and v.company_name else (v.name if v else vid[:8])
        )
        _write_row(ws, i, cols, [name, slot["count"], slot["amount"]])
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_cost_by_service_sheet(ws_wb, invoices, workorders_by_id):
    ws = ws_wb.create_sheet("Cost by Service")
    _write_title(ws, "Cost by Service Type", "Invoice totals grouped by the service type on the underlying work order.")

    by_service = defaultdict(
        lambda: {
            "invoice_count": 0,
            "total": 0.0,
            "approved": 0.0,
            "pending": 0.0,
            "rejected": 0.0,
            "paid": 0.0,
        }
    )
    for inv in invoices:
        wo = workorders_by_id.get(inv.work_order_id)
        service_name = "Unknown"
        if wo:
            st = wo.service_type
            if hasattr(st, "service"):
                service_name = st.service
            elif st:
                service_name = str(st)
        amount = float(inv.total_amount or 0)
        slot = by_service[service_name]
        slot["invoice_count"] += 1
        slot["total"] += amount
        status = _enum_value(inv.invoice_status)
        if status == "APPROVED":
            slot["approved"] += amount
        elif status == "PENDING":
            slot["pending"] += amount
        elif status == "REJECTED":
            slot["rejected"] += amount
        elif status == "PAID":
            slot["paid"] += amount

    header_row = 5
    cols = [
        ("Service", 26, None),
        ("Invoice count", 14, "#,##0"),
        ("Total ($)", 18, "$#,##0.00"),
        ("Approved ($)", 18, "$#,##0.00"),
        ("Pending ($)", 18, "$#,##0.00"),
        ("Rejected ($)", 18, "$#,##0.00"),
        ("Paid ($)", 18, "$#,##0.00"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    rows = sorted(by_service.items(), key=lambda kv: kv[1]["total"], reverse=True)
    for i, (service, slot) in enumerate(rows, start=header_row + 1):
        _write_row(
            ws,
            i,
            cols,
            [
                service,
                slot["invoice_count"],
                slot["total"],
                slot["approved"],
                slot["pending"],
                slot["rejected"],
                slot["paid"],
            ],
        )
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_wo_status_sheet(ws_wb, work_orders):
    ws = ws_wb.create_sheet("WO by Status")
    _write_title(ws, "Work Orders by Status", "Count of work orders in each lifecycle status.")

    by_status = defaultdict(int)
    for w in work_orders:
        by_status[_enum_value(getattr(w, "current_status", None)) or "UNKNOWN"] += 1

    header_row = 5
    cols = [("Status", 18, None), ("Count", 12, "#,##0")]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    rows = sorted(by_status.items(), key=lambda kv: kv[1], reverse=True)
    for i, (status, count) in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, cols, [status, count])
    _finalize_data_sheet(ws, header_row, cols, len(rows))


def _write_msa_expiry_sheet(ws_wb, msas, vendors_by_id, work_orders):
    ws = ws_wb.create_sheet("MSAs Expiring (90d)")
    _write_title(ws, "MSAs Expiring in the Next 90 Days", "Includes vendor name and current active WO count for context.")

    today = datetime.utcnow().date()
    cutoff = today + timedelta(days=90)
    active_wo_by_vendor = defaultdict(int)
    for w in work_orders:
        if _enum_value(getattr(w, "current_status", None)) in ("CANCELLED", "CLOSED"):
            continue
        if w.assigned_vendor:
            active_wo_by_vendor[w.assigned_vendor] += 1

    rows = []
    for m in msas:
        if not m.expiration_date:
            continue
        if not (today <= m.expiration_date <= cutoff):
            continue
        v = vendors_by_id.get(m.vendor_id)
        rows.append(
            [
                v.company_name if v and v.company_name else (v.name if v else "Unknown"),
                m.version or "",
                m.expiration_date,
                (m.expiration_date - today).days,
                m.status or "",
                active_wo_by_vendor.get(m.vendor_id, 0),
            ]
        )
    rows.sort(key=lambda r: r[3])

    header_row = 5
    cols = [
        ("Vendor", 30, None),
        ("Version", 10, None),
        ("Expires", 14, "yyyy-mm-dd"),
        ("Days left", 12, "#,##0"),
        ("Status", 14, None),
        ("Active WOs", 12, "#,##0"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)
    for i, row in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, cols, row)
    _finalize_data_sheet(ws, header_row, cols, len(rows))


# =====================================================================
# Invoices workbook — executive rebuild
# =====================================================================
#
# Tab order (Executive Summary first, set as active on save):
#   1. Executive Summary  — KPIs, status mix, top vendors, watch list
#   2. Invoices Data      — flat source with derived days-to-pay columns
#   3. Vendor Scorecard   — per-vendor formulas vs flat
#   4. Aging              — past-due aging buckets
#   5. Line Items         — flat line item list
#   6. Throughput by Hour — created/approved/paid counts by hour-of-day
#   7. Data Quality       — coverage / null counts / duplicates
#
# All scalar KPIs and per-vendor metrics are Excel formulas pointing at
# the flat tab so editing rows in place recalculates the whole book.

I_FLAT_SHEET = "Invoices Data"
I_REPORT_TITLE = "Invoices Report"
I_HEADER_ROW = 4


def build_invoices_workbook(start=None, end=None):
    """Executive-quality invoices workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    invoices = db.session.execute(select(Invoice)).scalars().all()
    if start or end:
        invoices = [i for i in invoices if _in_range(i.invoice_date, start, end)]
    vendors_by_id = {v.id: v for v in db.session.execute(select(Vendor)).scalars().all()}
    workorders_by_id = {w.id: w for w in db.session.execute(select(WorkOrder)).scalars().all()}
    line_items = db.session.execute(select(LineItem)).scalars().all()
    line_items_by_invoice = defaultdict(list)
    for li in line_items:
        line_items_by_invoice[li.invoice_id].append(li)

    period = _period_label(start, end)

    flat_last_row = _i_write_flat(wb, invoices, vendors_by_id, workorders_by_id)
    _i_write_exec_summary(wb, invoices, vendors_by_id, period, flat_last_row)
    _i_write_vendor_scorecard(wb, invoices, vendors_by_id, flat_last_row)
    _i_write_aging(wb, invoices)
    _i_write_line_items(wb, invoices, vendors_by_id, line_items_by_invoice)
    _i_write_throughput_by_hour(wb, invoices)
    _i_write_data_quality(wb, invoices)

    exec_idx = wb.sheetnames.index("Executive Summary")
    if exec_idx != 0:
        wb.move_sheet("Executive Summary", offset=-exec_idx)
    wb.active = 0

    exec_ws = wb["Executive Summary"]
    exec_ws.sheet_view.zoomScale = 100
    exec_ws.sheet_view.selection[0].activeCell = "A1"
    exec_ws.sheet_view.selection[0].sqref = "A1"

    for ws in wb.worksheets:
        if ws.title == "Executive Summary":
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            _wo_print_setup(ws, period, ws.title, gridlines=False, report_title=I_REPORT_TITLE)
        else:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            _wo_print_setup(ws, period, ws.title, gridlines=True, report_title=I_REPORT_TITLE)

    return _to_workbook_bytes(wb)


# ---------------------------------------------------------------------
# Invoices Data (flat) — source of truth
# ---------------------------------------------------------------------

def _i_write_flat(wb, invoices, vendors_by_id, workorders_by_id):
    ws = wb.create_sheet(I_FLAT_SHEET)
    note = (
        "Source-of-truth flat list. Columns L-O are derived in-cell: "
        "Days to Approval, Days to Pay, Past Due?, Days Overdue."
    )
    ws.cell(row=1, column=1, value="Invoices - source data").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Invoice ID", 14, None),
        ("Vendor", 28, None),
        ("WO #", 10, FMT_INT),
        ("Status", 14, None),
        ("Total Amount", 14, FMT_CURRENCY),
        ("Invoice Date", 14, FMT_DATE),
        ("Due Date", 14, FMT_DATE),
        ("Approved At", 16, FMT_DATETIME),
        ("Paid At", 16, FMT_DATETIME),
        ("Rejected At", 16, FMT_DATETIME),
        ("Created", 16, FMT_DATETIME),
        ("Days to Approval", 16, FMT_HOURS),
        ("Days to Pay", 14, FMT_HOURS),
        ("Past Due?", 10, None),
        ("Days Overdue", 14, FMT_INT),
    ]
    _apply_widths(ws, cols)

    header_row = I_HEADER_ROW
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    sorted_invoices = sorted(
        invoices,
        key=lambda i: -((i.invoice_date or datetime.min).timestamp() if i.invoice_date else 0),
    )
    for offset, inv in enumerate(sorted_invoices):
        row = header_row + 1 + offset
        v = vendors_by_id.get(inv.vendor_id) if inv.vendor_id else None
        wo = workorders_by_id.get(inv.work_order_id) if inv.work_order_id else None
        wo_code = getattr(wo, "work_order_code", None) if wo else None

        # F = Invoice Date, G = Due Date, H = Approved At, I = Paid At, D = Status
        days_to_approval = (
            f'=IF(AND(F{row}<>"",H{row}<>""),H{row}-F{row},"")'
        )
        days_to_pay = (
            f'=IF(AND(H{row}<>"",I{row}<>""),I{row}-H{row},"")'
        )
        past_due = (
            f'=IF(AND(G{row}<>"",I{row}="",D{row}<>"PAID",TODAY()>G{row}),"Yes","No")'
        )
        days_overdue = (
            f'=IF(N{row}="Yes",TODAY()-G{row},"")'
        )

        values = [
            inv.id[:8] if inv.id else "",
            _vendor_name(v),
            wo_code,
            _enum_value(getattr(inv, "invoice_status", None)),
            float(getattr(inv, "total_amount", 0) or 0),
            _naive(getattr(inv, "invoice_date", None)),
            _naive(getattr(inv, "due_date", None)),
            _naive(getattr(inv, "approved_at", None)),
            _naive(getattr(inv, "paid_at", None)),
            _naive(getattr(inv, "rejected_at", None)),
            _naive(getattr(inv, "created_at", None)),
            days_to_approval,
            days_to_pay,
            past_due,
            days_overdue,
        ]

        for c_idx, ((_, _w, fmt), value) in enumerate(zip(cols, values), start=1):
            cell = ws.cell(row=row, column=c_idx, value=value)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_data_row = header_row + len(sorted_invoices)
    last_col_letter = get_column_letter(len(cols))

    if sorted_invoices:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        rng = f"D{header_row + 1}:D{last_data_row}"
        for status, fill in (
            ("PAID", WO_GREEN_FILL),
            ("APPROVED", WO_GREEN_FILL),
            ("SUBMITTED", WO_AMBER_FILL),
            ("DRAFT", WO_SLATE_FILL),
            ("REJECTED", WO_RED_FILL),
        ):
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill)
            )
        _wo_data_bar(ws, "E", header_row + 1, last_data_row)
        ws.conditional_formatting.add(
            f"N{header_row + 1}:N{last_data_row}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=WO_RED_FILL),
        )

    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, last_col_letter, max(last_data_row, header_row))
    return last_data_row


# ---------------------------------------------------------------------
# Executive Summary
# ---------------------------------------------------------------------

def _i_write_exec_summary(wb, invoices, vendors_by_id, period, flat_last_row):
    ws = wb.create_sheet("Executive Summary")
    flat = _wo_quoted(I_FLAT_SHEET)
    n = len(invoices)
    a, b = I_HEADER_ROW + 1, flat_last_row

    ws.merge_cells("A1:F1")
    ws["A1"] = I_REPORT_TITLE
    ws["A1"].font = WO_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"] = f"Period: {period}    ·    {n:,} invoice(s)"
    ws["A2"].font = WO_NOTE_FONT
    ws["A3"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"].font = WO_NOTE_FONT

    kpi_row = 6
    kpi_specs = [
        ("Total invoiced",
         f"=SUM({flat}!E{a}:E{b})", FMT_CURRENCY),
        ("Invoice count",
         f"=COUNTA({flat}!A{a}:A{b})", FMT_INT),
        ("Approval rate",
         f'=IFERROR((COUNTIF({flat}!D{a}:D{b},"APPROVED")'
         f'+COUNTIF({flat}!D{a}:D{b},"PAID"))'
         f'/COUNTA({flat}!D{a}:D{b}),0)', FMT_PCT),
        ("Average days to approval",
         f'=IFERROR(AVERAGE({flat}!L{a}:L{b}),0)', "#,##0.0"),
        ("Average days to pay",
         f'=IFERROR(AVERAGE({flat}!M{a}:M{b}),0)', "#,##0.0"),
        ("% past due",
         f'=IFERROR(COUNTIF({flat}!N{a}:N{b},"Yes")'
         f'/COUNTA({flat}!N{a}:N{b}),0)', FMT_PCT),
    ]
    for idx, (label, formula, fmt) in enumerate(kpi_specs):
        col = (idx % 3) * 2 + 1
        row = kpi_row + (idx // 3) * 3
        l = ws.cell(row=row, column=col, value=label); l.font = WO_KPI_LABEL_FONT
        v = ws.cell(row=row + 1, column=col, value=formula)
        v.font = WO_KPI_VALUE_FONT
        v.number_format = fmt
        for c in (l, v):
            c.border = THIN_BORDER

    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 22

    section_row = kpi_row + 7
    ws.cell(row=section_row, column=1, value="STATUS BREAKDOWN").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    cols = [("Status", 22, None), ("Count", 12, FMT_INT), ("Total", 16, FMT_CURRENCY), ("Share", 10, FMT_PCT)]
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    statuses = ["DRAFT", "SUBMITTED", "APPROVED", "REJECTED", "PAID"]
    total_formula = f"COUNTA({flat}!D{a}:D{b})"
    for i, status in enumerate(statuses):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=status).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({flat}!D{a}:D{b},"{status}")').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIF({flat}!D{a}:D{b},"{status}",{flat}!E{a}:E{b})'
                ).number_format = FMT_CURRENCY
        ws.cell(row=r, column=4,
                value=f'=IFERROR(COUNTIF({flat}!D{a}:D{b},"{status}")/{total_formula},0)'
                ).number_format = FMT_PCT
    status_last = section_row + len(statuses)

    section_row = status_last + 3
    ws.cell(row=section_row, column=1, value="TOP 5 VENDORS BY INVOICED $").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    for c_idx, label in enumerate(["Vendor", "Invoices", "Invoiced $"], start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    by_vendor = defaultdict(lambda: {"count": 0, "total": 0.0, "name": ""})
    for inv in invoices:
        if not inv.vendor_id:
            continue
        slot = by_vendor[inv.vendor_id]
        slot["count"] += 1
        slot["total"] += float(getattr(inv, "total_amount", 0) or 0)
        slot["name"] = _vendor_name(vendors_by_id.get(inv.vendor_id))
    top = sorted(by_vendor.values(), key=lambda x: x["total"], reverse=True)[:5]
    for i, slot in enumerate(top):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=slot["name"]).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=slot["count"]).number_format = FMT_INT
        ws.cell(row=r, column=3, value=slot["total"]).number_format = FMT_CURRENCY
    top_last = section_row + max(len(top), 1)

    section_row = top_last + 3
    ws.cell(row=section_row, column=1, value="WATCH LIST").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    for c_idx, label in enumerate(["Indicator", "Count"], start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    watch_specs = [
        ("Past due invoices",
         f'=COUNTIF({flat}!N{a}:N{b},"Yes")'),
        ("Rejected invoices",
         f'=COUNTIF({flat}!D{a}:D{b},"REJECTED")'),
        ("Submitted (awaiting approval)",
         f'=COUNTIF({flat}!D{a}:D{b},"SUBMITTED")'),
        ("Days overdue >= 30 (any)",
         f'=COUNTIFS({flat}!O{a}:O{b},">=30")'),
        ("Days overdue >= 60 (any)",
         f'=COUNTIFS({flat}!O{a}:O{b},">=60")'),
    ]
    for i, (label, formula) in enumerate(watch_specs):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=formula).number_format = FMT_INT


# ---------------------------------------------------------------------
# Vendor Scorecard
# ---------------------------------------------------------------------

def _i_write_vendor_scorecard(wb, invoices, vendors_by_id, flat_last_row):
    ws = wb.create_sheet("Vendor Scorecard")
    flat = _wo_quoted(I_FLAT_SHEET)
    note = (
        "Per-vendor performance vs the flat data tab. All metrics are formulas; "
        "edit rows in 'Invoices Data' and this tab recalculates."
    )
    ws.cell(row=1, column=1, value="Vendor Scorecard").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Vendor", 30, None),
        ("Invoices", 10, FMT_INT),
        ("Total Invoiced", 16, FMT_CURRENCY),
        ("Approval Rate", 14, FMT_PCT),
        ("Rejection Rate", 14, FMT_PCT),
        ("Avg Days to Approval", 18, "#,##0.0"),
        ("Avg Days to Pay", 16, "#,##0.0"),
        ("Past Due Count", 14, FMT_INT),
        ("% Past Due", 12, FMT_PCT),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    vendor_names = sorted(
        {_vendor_name(vendors_by_id.get(inv.vendor_id))
         for inv in invoices if inv.vendor_id}
    )
    a, b = I_HEADER_ROW + 1, flat_last_row
    for i, name in enumerate(vendor_names):
        r = header_row + 1 + i
        v_ref = f"A{r}"
        ws.cell(row=r, column=1, value=name).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({flat}!B{a}:B{b},{v_ref})').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIF({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b})'
                ).number_format = FMT_CURRENCY
        ws.cell(row=r, column=4,
                value=(
                    f'=IFERROR(('
                    f'COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!D{a}:D{b},"APPROVED")'
                    f'+COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!D{a}:D{b},"PAID"))'
                    f'/COUNTIF({flat}!B{a}:B{b},{v_ref}),0)'
                )).number_format = FMT_PCT
        ws.cell(row=r, column=5,
                value=f'=IFERROR(COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!D{a}:D{b},"REJECTED")/COUNTIF({flat}!B{a}:B{b},{v_ref}),0)'
                ).number_format = FMT_PCT
        ws.cell(row=r, column=6,
                value=f'=IFERROR(AVERAGEIFS({flat}!L{a}:L{b},{flat}!B{a}:B{b},{v_ref},{flat}!L{a}:L{b},">=0"),0)'
                ).number_format = "#,##0.0"
        ws.cell(row=r, column=7,
                value=f'=IFERROR(AVERAGEIFS({flat}!M{a}:M{b},{flat}!B{a}:B{b},{v_ref},{flat}!M{a}:M{b},">=0"),0)'
                ).number_format = "#,##0.0"
        ws.cell(row=r, column=8,
                value=f'=COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!N{a}:N{b},"Yes")'
                ).number_format = FMT_INT
        ws.cell(row=r, column=9,
                value=f'=IFERROR(COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!N{a}:N{b},"Yes")/COUNTIF({flat}!B{a}:B{b},{v_ref}),0)'
                ).number_format = FMT_PCT
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    last_row = header_row + len(vendor_names)
    if vendor_names:
        ws.auto_filter.ref = f"A{header_row}:I{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        # Highlight rejection rate >10%
        ws.conditional_formatting.add(
            f"E{header_row + 1}:E{last_row}",
            CellIsRule(operator="greaterThan", formula=["0.1"], fill=WO_RED_FILL),
        )
        ws.conditional_formatting.add(
            f"I{header_row + 1}:I{last_row}",
            CellIsRule(operator="greaterThan", formula=["0.2"], fill=WO_RED_FILL),
        )
        _wo_data_bar(ws, "C", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "I", max(last_row, header_row))


# ---------------------------------------------------------------------
# Aging
# ---------------------------------------------------------------------

def _i_write_aging(wb, invoices):
    ws = wb.create_sheet("Aging")
    note = (
        "Aging buckets for unpaid, past-due invoices, calculated from "
        "today's date when the workbook was generated."
    )
    ws.cell(row=1, column=1, value="Invoice Aging").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Bucket", 18, None),
        ("Invoices", 10, FMT_INT),
        ("Total $", 16, FMT_CURRENCY),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    today = datetime.utcnow().date()
    buckets = {
        "Current (not past due)": (None, 0, [], 0.0),
        "1-30 days": (1, 30, [], 0.0),
        "31-60 days": (31, 60, [], 0.0),
        "61-90 days": (61, 90, [], 0.0),
        "90+ days": (91, None, [], 0.0),
    }
    for inv in invoices:
        if getattr(inv, "paid_at", None):
            continue
        if _enum_value(getattr(inv, "invoice_status", None)) == "PAID":
            continue
        due = getattr(inv, "due_date", None)
        if not due:
            continue
        due_d = due.date() if hasattr(due, "date") else due
        days = (today - due_d).days
        amt = float(getattr(inv, "total_amount", 0) or 0)
        if days <= 0:
            buckets["Current (not past due)"][2].append(inv)
            buckets["Current (not past due)"] = (None, 0,
                                                 buckets["Current (not past due)"][2],
                                                 buckets["Current (not past due)"][3] + amt)
        elif days <= 30:
            buckets["1-30 days"][2].append(inv)
            buckets["1-30 days"] = (1, 30, buckets["1-30 days"][2], buckets["1-30 days"][3] + amt)
        elif days <= 60:
            buckets["31-60 days"][2].append(inv)
            buckets["31-60 days"] = (31, 60, buckets["31-60 days"][2], buckets["31-60 days"][3] + amt)
        elif days <= 90:
            buckets["61-90 days"][2].append(inv)
            buckets["61-90 days"] = (61, 90, buckets["61-90 days"][2], buckets["61-90 days"][3] + amt)
        else:
            buckets["90+ days"][2].append(inv)
            buckets["90+ days"] = (91, None, buckets["90+ days"][2], buckets["90+ days"][3] + amt)

    row = header_row
    for label, (_lo, _hi, items, total) in buckets.items():
        row += 1
        ws.cell(row=row, column=1, value=label).font = WO_BODY_FONT
        ws.cell(row=row, column=2, value=len(items)).number_format = FMT_INT
        ws.cell(row=row, column=3, value=total).number_format = FMT_CURRENCY
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=row, column=c_idx).border = THIN_BORDER

    if buckets:
        _wo_data_bar(ws, "C", header_row + 1, row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "C", row)


# ---------------------------------------------------------------------
# Line Items
# ---------------------------------------------------------------------

def _i_write_line_items(wb, invoices, vendors_by_id, line_items_by_invoice):
    ws = wb.create_sheet("Line Items")
    note = "Every line item across every invoice in the period."
    ws.cell(row=1, column=1, value="Line Items").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Invoice ID", 14, None),
        ("Vendor", 28, None),
        ("Description", 38, None),
        ("Quantity", 10, "#,##0.00"),
        ("Unit Price", 12, FMT_CURRENCY_DEC),
        ("Amount", 14, FMT_CURRENCY_DEC),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    rows = []
    for inv in invoices:
        items = line_items_by_invoice.get(inv.id, [])
        v = vendors_by_id.get(inv.vendor_id) if inv.vendor_id else None
        for li in items:
            qty = float(getattr(li, "quantity", 0) or 0)
            unit_price = float(getattr(li, "unit_price", 0) or 0)
            amount = float(getattr(li, "amount", None) or qty * unit_price)
            rows.append([
                inv.id[:8] if inv.id else "",
                _vendor_name(v),
                getattr(li, "description", None) or "",
                qty,
                unit_price,
                amount,
            ])

    for i, vals in enumerate(rows):
        r = header_row + 1 + i
        for c_idx, ((_, _w, fmt), val) in enumerate(zip(cols, vals), start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = header_row + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:F{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "F", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "F", max(last_row, header_row))


# ---------------------------------------------------------------------
# Throughput by Hour
# ---------------------------------------------------------------------

def _i_write_throughput_by_hour(wb, invoices):
    ws = wb.create_sheet("Throughput by Hour")
    note = "Created, approved, and paid counts per hour-of-day."
    ws.cell(row=1, column=1, value="Throughput by Hour-of-Day").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Hour (UTC)", 10, "00"),
        ("Created", 10, FMT_INT),
        ("Approved", 10, FMT_INT),
        ("Paid", 10, FMT_INT),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    created_by_h = defaultdict(int)
    approved_by_h = defaultdict(int)
    paid_by_h = defaultdict(int)
    for inv in invoices:
        if getattr(inv, "created_at", None):
            created_by_h[inv.created_at.hour] += 1
        if getattr(inv, "approved_at", None):
            approved_by_h[inv.approved_at.hour] += 1
        if getattr(inv, "paid_at", None):
            paid_by_h[inv.paid_at.hour] += 1

    for h in range(24):
        r = header_row + 1 + h
        cells = [
            (1, h, "00"),
            (2, created_by_h.get(h, 0), FMT_INT),
            (3, approved_by_h.get(h, 0), FMT_INT),
            (4, paid_by_h.get(h, 0), FMT_INT),
        ]
        for col, val, fmt in cells:
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = header_row + 24
    ws.auto_filter.ref = f"A{header_row}:D{last_row}"
    ws.freeze_panes = ws[f"A{header_row + 1}"]
    _wo_data_bar(ws, "B", header_row + 1, last_row)
    _wo_data_bar(ws, "C", header_row + 1, last_row, color="64748B")
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "D", last_row)


# ---------------------------------------------------------------------
# Data Quality
# ---------------------------------------------------------------------

def _i_write_data_quality(wb, invoices):
    ws = wb.create_sheet("Data Quality")
    note = "Coverage and completeness of the invoice source data."
    ws.cell(row=1, column=1, value="Data Quality").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    n = len(invoices)
    miss_invoice_date = sum(1 for inv in invoices if not getattr(inv, "invoice_date", None))
    miss_due = sum(1 for inv in invoices if not getattr(inv, "due_date", None))
    miss_amount = sum(1 for inv in invoices if not getattr(inv, "total_amount", None))
    miss_vendor = sum(1 for inv in invoices if not inv.vendor_id)
    miss_wo = sum(1 for inv in invoices if not inv.work_order_id)

    seen = set()
    dup_ids = 0
    for inv in invoices:
        if inv.id is None:
            continue
        if inv.id in seen:
            dup_ids += 1
        else:
            seen.add(inv.id)

    rows = [
        ("Total invoice rows", n, FMT_INT),
        ("Missing Invoice Date", miss_invoice_date, FMT_INT),
        ("Missing Due Date", miss_due, FMT_INT),
        ("Missing Total Amount", miss_amount, FMT_INT),
        ("Missing Vendor", miss_vendor, FMT_INT),
        ("Missing Work Order link", miss_wo, FMT_INT),
        ("Duplicate invoice ids", dup_ids, FMT_INT),
        ("% missing invoice_date", (miss_invoice_date / n) if n else 0, FMT_PCT),
        ("% missing due_date", (miss_due / n) if n else 0, FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.font = WO_KPI_VALUE_FONT
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.print_title_rows = "1:4"
    _wo_set_print_area(ws, "B", 4 + len(rows))


# =====================================================================
# Tickets workbook — executive rebuild
# =====================================================================
#
# Tab order (Executive Summary first, set as active on save):
#   1. Executive Summary       — KPIs, status mix, top vendors, watch list
#   2. Tickets Data            — flat source with derived duration columns
#   3. Vendor Scorecard        — per-vendor formulas vs flat
#   4. Contractor Utilization  — per-contractor minutes worked + jobs
#   5. Service Benchmarks      — per-service throughput + duration quartiles
#   6. Duration Analysis       — overall quartiles + outliers
#   7. Throughput by Hour      — created/approved counts by hour-of-day
#   8. Data Quality            — coverage / null counts / duplicates
#
# All scalar KPIs and per-vendor metrics are Excel formulas pointing at
# the flat tab so editing rows in place recalculates the whole book.

T_FLAT_SHEET = "Tickets Data"
T_REPORT_TITLE = "Tickets Report"
T_HEADER_ROW = 4


def build_tickets_workbook(start=None, end=None):
    """Executive-quality tickets workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    tickets = db.session.execute(select(Ticket)).scalars().all()
    if start or end:
        tickets = [t for t in tickets if _in_range(t.created_at, start, end)]
    vendors_by_id = {v.id: v for v in db.session.execute(select(Vendor)).scalars().all()}
    workorders_by_id = {w.id: w for w in db.session.execute(select(WorkOrder)).scalars().all()}

    period = _period_label(start, end)

    flat_last_row = _t_write_flat(wb, tickets, vendors_by_id, workorders_by_id)
    _t_write_exec_summary(wb, tickets, vendors_by_id, period, flat_last_row)
    _t_write_vendor_scorecard(wb, tickets, vendors_by_id, flat_last_row)
    _t_write_contractor_utilization(wb, tickets)
    _t_write_service_benchmarks(wb, tickets)
    _t_write_duration_analysis(wb, tickets)
    _t_write_throughput_by_hour(wb, tickets)
    _t_write_data_quality(wb, tickets)

    exec_idx = wb.sheetnames.index("Executive Summary")
    if exec_idx != 0:
        wb.move_sheet("Executive Summary", offset=-exec_idx)
    wb.active = 0

    exec_ws = wb["Executive Summary"]
    exec_ws.sheet_view.zoomScale = 100
    exec_ws.sheet_view.selection[0].activeCell = "A1"
    exec_ws.sheet_view.selection[0].sqref = "A1"

    for ws in wb.worksheets:
        if ws.title == "Executive Summary":
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            _wo_print_setup(ws, period, ws.title, gridlines=False, report_title=T_REPORT_TITLE)
        else:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            _wo_print_setup(ws, period, ws.title, gridlines=True, report_title=T_REPORT_TITLE)

    return _to_workbook_bytes(wb)


# ---------------------------------------------------------------------
# Tickets Data (flat) — source of truth
# ---------------------------------------------------------------------

def _t_write_flat(wb, tickets, vendors_by_id, workorders_by_id):
    ws = wb.create_sheet(T_FLAT_SHEET)
    note = (
        "Source-of-truth flat list. Columns R-V are derived in-cell from "
        "the timestamps to the left (durations in hours, time to approval, "
        "on-time, open-flag)."
    )
    ws.cell(row=1, column=1, value="Tickets - source data").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("WO #", 10, FMT_INT),
        ("Vendor", 28, None),
        ("Ticket ID", 12, None),
        ("Description", 38, None),
        ("Status", 18, None),
        ("Priority", 10, None),
        ("Service", 18, None),
        ("Assigned Contractor", 22, None),
        ("Start Time", 16, FMT_DATETIME),
        ("End Time", 16, FMT_DATETIME),
        ("Due Date", 12, FMT_DATE),
        ("Created", 16, FMT_DATETIME),
        ("Approved At", 16, FMT_DATETIME),
        ("Rejected At", 16, FMT_DATETIME),
        ("Est Quantity", 12, "#,##0.00"),
        ("Unit", 8, None),
        ("Anomaly?", 10, None),
        ("Duration (hrs)", 14, FMT_HOURS),
        ("Time to Approval (hrs)", 16, FMT_HOURS),
        ("Time to Completion (hrs)", 18, FMT_HOURS),
        ("On Time?", 10, None),
        ("Open?", 8, None),
    ]
    _apply_widths(ws, cols)

    header_row = T_HEADER_ROW
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    sorted_tickets = sorted(
        tickets,
        key=lambda t: -((t.created_at or datetime.min).timestamp() if t.created_at else 0),
    )
    for offset, t in enumerate(sorted_tickets):
        row = header_row + 1 + offset
        wo = workorders_by_id.get(t.work_order_id) if t.work_order_id else None
        wo_code = getattr(wo, "work_order_code", None) if wo else None
        v = vendors_by_id.get(t.vendor_id) if t.vendor_id else None
        service = t.service.service if getattr(t, "service", None) else None

        # I = Start (col 9), J = End (col 10), K = Due (col 11),
        # L = Created (col 12), M = Approved (col 13), E = Status (col 5)
        duration = f'=IF(AND(I{row}<>"",J{row}<>""),(J{row}-I{row})*24,"")'
        time_to_approval = f'=IF(AND(L{row}<>"",M{row}<>""),(M{row}-L{row})*24,"")'
        time_to_completion = f'=IF(AND(L{row}<>"",J{row}<>""),(J{row}-L{row})*24,"")'
        on_time = (
            f'=IF(AND(J{row}<>"",K{row}<>""),'
            f'IF(J{row}<=K{row},"Yes","No"),"")'
        )
        open_flag = (
            f'=IF(OR(E{row}="UNASSIGNED",E{row}="ASSIGNED",E{row}="IN_PROGRESS"),'
            f'"Yes","No")'
        )

        values = [
            wo_code,
            _vendor_name(v),
            t.id[:8] if t.id else "",
            t.description or "",
            _enum_value(getattr(t, "status", None)),
            _enum_value(getattr(t, "priority", None)),
            service or "",
            getattr(t, "assigned_contractor", None) or "",
            _naive(getattr(t, "start_time", None)),
            _naive(getattr(t, "end_time", None)),
            _naive(getattr(t, "due_date", None)),
            _naive(t.created_at),
            _naive(getattr(t, "approved_at", None)),
            _naive(getattr(t, "rejected_at", None)),
            float(getattr(t, "estimated_quantity", 0) or 0),
            getattr(t, "unit", None) or "",
            "Yes" if getattr(t, "anomaly_flag", False) else "No",
            duration,
            time_to_approval,
            time_to_completion,
            on_time,
            open_flag,
        ]

        for c_idx, ((_, _w, fmt), value) in enumerate(zip(cols, values), start=1):
            cell = ws.cell(row=row, column=c_idx, value=value)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_data_row = header_row + len(sorted_tickets)
    last_col = len(cols)
    last_col_letter = get_column_letter(last_col)

    if sorted_tickets:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        rng = f"E{header_row + 1}:E{last_data_row}"
        for status, fill in (
            ("APPROVED", WO_GREEN_FILL),
            ("COMPLETED", WO_GREEN_FILL),
            ("PENDING_APPROVAL", WO_AMBER_FILL),
            ("IN_PROGRESS", WO_AMBER_FILL),
            ("ASSIGNED", WO_AMBER_FILL),
            ("UNASSIGNED", WO_SLATE_FILL),
            ("REJECTED", WO_RED_FILL),
        ):
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill)
            )
        _wo_data_bar(ws, "R", header_row + 1, last_data_row, color="334155")
        _wo_data_bar(ws, "S", header_row + 1, last_data_row)
        ws.conditional_formatting.add(
            f"U{header_row + 1}:U{last_data_row}",
            CellIsRule(operator="equal", formula=['"No"'], fill=WO_RED_FILL),
        )

    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, last_col_letter, max(last_data_row, header_row))
    return last_data_row


# ---------------------------------------------------------------------
# Executive Summary
# ---------------------------------------------------------------------

def _t_write_exec_summary(wb, tickets, vendors_by_id, period, flat_last_row):
    ws = wb.create_sheet("Executive Summary")
    flat = _wo_quoted(T_FLAT_SHEET)
    n = len(tickets)
    a, b = T_HEADER_ROW + 1, flat_last_row

    ws.merge_cells("A1:F1")
    ws["A1"] = T_REPORT_TITLE
    ws["A1"].font = WO_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"] = f"Period: {period}    ·    {n:,} ticket(s)"
    ws["A2"].font = WO_NOTE_FONT
    ws["A3"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"].font = WO_NOTE_FONT

    kpi_row = 6
    kpi_specs = [
        ("Total tickets", f"=COUNTA({flat}!C{a}:C{b})", FMT_INT),
        ("Average duration (hrs)",
         f'=IFERROR(AVERAGE({flat}!R{a}:R{b}),0)', FMT_HOURS),
        ("On-time completion rate",
         f'=IFERROR(COUNTIF({flat}!U{a}:U{b},"Yes")'
         f'/(COUNTIF({flat}!U{a}:U{b},"Yes")'
         f'+COUNTIF({flat}!U{a}:U{b},"No")),0)', FMT_PCT),
        ("Open tickets", f'=COUNTIF({flat}!V{a}:V{b},"Yes")', FMT_INT),
        ("Average time to approval (hrs)",
         f'=IFERROR(AVERAGE({flat}!S{a}:S{b}),0)', FMT_HOURS),
        ("Pending approval count",
         f'=COUNTIF({flat}!E{a}:E{b},"PENDING_APPROVAL")', FMT_INT),
    ]
    for idx, (label, formula, fmt) in enumerate(kpi_specs):
        col = (idx % 3) * 2 + 1
        row = kpi_row + (idx // 3) * 3
        l = ws.cell(row=row, column=col, value=label); l.font = WO_KPI_LABEL_FONT
        v = ws.cell(row=row + 1, column=col, value=formula)
        v.font = WO_KPI_VALUE_FONT
        v.number_format = fmt
        for c in (l, v):
            c.border = THIN_BORDER

    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 22

    section_row = kpi_row + 7
    ws.cell(row=section_row, column=1, value="STATUS BREAKDOWN").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    cols = [("Status", 22, None), ("Count", 12, FMT_INT), ("Share", 10, FMT_PCT)]
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    statuses = [
        "UNASSIGNED", "ASSIGNED", "IN_PROGRESS",
        "PENDING_APPROVAL", "APPROVED", "REJECTED", "COMPLETED",
    ]
    total_formula = f"COUNTA({flat}!E{a}:E{b})"
    for i, status in enumerate(statuses):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=status).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({flat}!E{a}:E{b},"{status}")').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=IFERROR(COUNTIF({flat}!E{a}:E{b},"{status}")/{total_formula},0)'
                ).number_format = FMT_PCT
    status_last = section_row + len(statuses)

    section_row = status_last + 3
    ws.cell(row=section_row, column=1, value="TOP 5 VENDORS BY TICKET COUNT").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    for c_idx, label in enumerate(["Vendor", "Tickets", "Avg Duration (hrs)"], start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    by_vendor = defaultdict(lambda: {"count": 0, "name": "", "durations": []})
    for t in tickets:
        if not t.vendor_id:
            continue
        slot = by_vendor[t.vendor_id]
        slot["count"] += 1
        slot["name"] = _vendor_name(vendors_by_id.get(t.vendor_id))
        st, et = getattr(t, "start_time", None), getattr(t, "end_time", None)
        if st and et:
            slot["durations"].append((et - st).total_seconds() / 3600.0)
    top = sorted(by_vendor.values(), key=lambda x: x["count"], reverse=True)[:5]
    for i, slot in enumerate(top):
        r = section_row + 1 + i
        avg_dur = sum(slot["durations"]) / len(slot["durations"]) if slot["durations"] else 0
        ws.cell(row=r, column=1, value=slot["name"]).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=slot["count"]).number_format = FMT_INT
        ws.cell(row=r, column=3, value=avg_dur).number_format = FMT_HOURS
    top_last = section_row + max(len(top), 1)

    section_row = top_last + 3
    ws.cell(row=section_row, column=1, value="WATCH LIST").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    for c_idx, label in enumerate(["Indicator", "Count"], start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD; h.fill = WO_HEADER_FILL
    watch_specs = [
        ("Open tickets (unassigned, assigned, in progress)",
         f'=COUNTIF({flat}!V{a}:V{b},"Yes")'),
        ("Pending approval", f'=COUNTIF({flat}!E{a}:E{b},"PENDING_APPROVAL")'),
        ("Rejected tickets", f'=COUNTIF({flat}!E{a}:E{b},"REJECTED")'),
        ("Late completions (On Time = No)",
         f'=COUNTIF({flat}!U{a}:U{b},"No")'),
        ("Anomaly-flagged tickets", f'=COUNTIF({flat}!Q{a}:Q{b},"Yes")'),
    ]
    for i, (label, formula) in enumerate(watch_specs):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=formula).number_format = FMT_INT


# ---------------------------------------------------------------------
# Vendor Scorecard
# ---------------------------------------------------------------------

def _t_write_vendor_scorecard(wb, tickets, vendors_by_id, flat_last_row):
    ws = wb.create_sheet("Vendor Scorecard")
    flat = _wo_quoted(T_FLAT_SHEET)
    note = (
        "Per-vendor performance vs the flat data tab. All metrics are formulas; "
        "edit rows in 'Tickets Data' and this tab recalculates."
    )
    ws.cell(row=1, column=1, value="Vendor Scorecard").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Vendor", 30, None),
        ("Tickets", 10, FMT_INT),
        ("Avg Duration (hrs)", 16, FMT_HOURS),
        ("Avg Time to Approval (hrs)", 18, FMT_HOURS),
        ("Completed", 10, FMT_INT),
        ("Completion Rate", 14, FMT_PCT),
        ("Distinct Contractors", 14, FMT_INT),
        ("Tickets per Contractor", 16, "#,##0.0"),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    by_vendor = defaultdict(lambda: {"name": "", "contractors": set()})
    for t in tickets:
        if not t.vendor_id:
            continue
        slot = by_vendor[t.vendor_id]
        slot["name"] = _vendor_name(vendors_by_id.get(t.vendor_id))
        c = (getattr(t, "assigned_contractor", None) or "").strip().lower()
        if c:
            slot["contractors"].add(c)

    vendor_rows = sorted(by_vendor.values(), key=lambda x: x["name"])
    a, b = T_HEADER_ROW + 1, flat_last_row
    for i, slot in enumerate(vendor_rows):
        r = header_row + 1 + i
        v_ref = f"A{r}"
        ws.cell(row=r, column=1, value=slot["name"]).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({flat}!B{a}:B{b},{v_ref})').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=IFERROR(AVERAGEIFS({flat}!R{a}:R{b},{flat}!B{a}:B{b},{v_ref},{flat}!R{a}:R{b},">0"),0)'
                ).number_format = FMT_HOURS
        ws.cell(row=r, column=4,
                value=f'=IFERROR(AVERAGEIFS({flat}!S{a}:S{b},{flat}!B{a}:B{b},{v_ref},{flat}!S{a}:S{b},">0"),0)'
                ).number_format = FMT_HOURS
        ws.cell(row=r, column=5,
                value=f'=COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"COMPLETED")'
                ).number_format = FMT_INT
        ws.cell(row=r, column=6,
                value=f'=IFERROR(COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"COMPLETED")/COUNTIF({flat}!B{a}:B{b},{v_ref}),0)'
                ).number_format = FMT_PCT
        ws.cell(row=r, column=7, value=len(slot["contractors"])).number_format = FMT_INT
        ws.cell(row=r, column=8,
                value=f'=IFERROR(B{r}/G{r},0)').number_format = "#,##0.0"
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    last_row = header_row + len(vendor_rows)
    if vendor_rows:
        ws.auto_filter.ref = f"A{header_row}:H{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "B", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "H", max(last_row, header_row))


# ---------------------------------------------------------------------
# Contractor Utilization
# ---------------------------------------------------------------------

def _t_write_contractor_utilization(wb, tickets):
    ws = wb.create_sheet("Contractor Utilization")
    note = (
        "Per contractor: tickets worked, total billable minutes "
        "(end_time - start_time), average minutes per ticket. Tickets with "
        "missing timestamps are counted but contribute zero minutes."
    )
    ws.cell(row=1, column=1, value="Contractor Utilization").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Contractor", 28, None),
        ("Tickets", 10, FMT_INT),
        ("Total Minutes", 14, "#,##0"),
        ("Total Hours", 14, FMT_HOURS),
        ("Avg Minutes/Ticket", 18, "#,##0.0"),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    by_c = defaultdict(lambda: {"count": 0, "minutes": 0.0})
    for t in tickets:
        c = (getattr(t, "assigned_contractor", None) or "").strip()
        if not c:
            continue
        slot = by_c[c]
        slot["count"] += 1
        st, et = getattr(t, "start_time", None), getattr(t, "end_time", None)
        if st and et:
            slot["minutes"] += (et - st).total_seconds() / 60.0

    rows = sorted(by_c.items(), key=lambda kv: kv[1]["minutes"], reverse=True)
    for i, (name, slot) in enumerate(rows):
        r = header_row + 1 + i
        avg = (slot["minutes"] / slot["count"]) if slot["count"] else 0
        ws.cell(row=r, column=1, value=name).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=slot["count"]).number_format = FMT_INT
        ws.cell(row=r, column=3, value=slot["minutes"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=slot["minutes"] / 60.0).number_format = FMT_HOURS
        ws.cell(row=r, column=5, value=avg).number_format = "#,##0.0"
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    last_row = header_row + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:E{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "C", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "E", max(last_row, header_row))


# ---------------------------------------------------------------------
# Service Benchmarks
# ---------------------------------------------------------------------

def _t_durations_for(filter_fn, tickets):
    out = []
    for t in tickets:
        if not filter_fn(t):
            continue
        s, e = getattr(t, "start_time", None), getattr(t, "end_time", None)
        if s and e:
            hrs = (e - s).total_seconds() / 3600.0
            if hrs > 0:
                out.append(hrs)
    return out


def _t_write_service_benchmarks(wb, tickets):
    ws = wb.create_sheet("Service Benchmarks")
    note = (
        "Per service type ticket counts and duration quartiles. "
        "Quartiles computed in Python from the snapshot."
    )
    ws.cell(row=1, column=1, value="Service Type Benchmarks").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Service", 24, None),
        ("Count", 10, FMT_INT),
        ("Avg Duration (hrs)", 16, FMT_HOURS),
        ("Min", 8, FMT_HOURS),
        ("p25", 8, FMT_HOURS),
        ("Median", 10, FMT_HOURS),
        ("p75", 8, FMT_HOURS),
        ("p90", 8, FMT_HOURS),
        ("Max", 10, FMT_HOURS),
        ("Outliers (>2x median)", 16, FMT_INT),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    services = sorted(
        {(t.service.service if getattr(t, "service", None) else "(no service)")
         for t in tickets}
    )
    row = header_row
    for svc in services:
        row += 1
        same = [t for t in tickets
                if (getattr(t.service, "service", None) if getattr(t, "service", None) else "(no service)") == svc]
        durs = _t_durations_for(
            lambda t, _svc=svc: ((getattr(t.service, "service", None) if getattr(t, "service", None) else "(no service)") == _svc),
            tickets,
        )
        avg = sum(durs) / len(durs) if durs else 0
        mn, p25, med, p75, p90, mx = _quartiles(durs)
        outliers = sum(1 for d in durs if med and d > 2 * med)
        values = [svc, len(same), avg, mn, p25, med, p75, p90, mx, outliers]
        for c_idx, ((_, _w, fmt), val) in enumerate(zip(cols, values), start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = row
    if services:
        ws.auto_filter.ref = f"A{header_row}:J{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "F", header_row + 1, last_row, color="334155")
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "J", max(last_row, header_row))


# ---------------------------------------------------------------------
# Duration Analysis
# ---------------------------------------------------------------------

def _t_write_duration_analysis(wb, tickets):
    ws = wb.create_sheet("Duration Analysis")
    note = (
        "Overall duration distribution across all tickets with both "
        "start_time and end_time. Outliers are jobs longer than 2x median."
    )
    ws.cell(row=1, column=1, value="Duration Analysis").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    durs = _t_durations_for(lambda t: True, tickets)
    mn, p25, med, p75, p90, mx = _quartiles(durs)
    avg = sum(durs) / len(durs) if durs else 0
    excluded = len(tickets) - len(durs)

    rows = [
        ("Tickets included (have start + end)", len(durs), FMT_INT),
        ("Tickets excluded (open or never started)", excluded, FMT_INT),
        ("Average duration (hrs)", avg, FMT_HOURS),
        ("Min", mn, FMT_HOURS),
        ("p25", p25, FMT_HOURS),
        ("Median", med, FMT_HOURS),
        ("p75", p75, FMT_HOURS),
        ("p90", p90, FMT_HOURS),
        ("Max", mx, FMT_HOURS),
        ("Outliers (>2x median)", sum(1 for d in durs if med and d > 2 * med), FMT_INT),
    ]
    for c_idx, (label, val, fmt) in enumerate(rows):
        r = 5 + c_idx
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        v = ws.cell(row=r, column=2, value=val)
        v.font = WO_KPI_VALUE_FONT
        v.number_format = fmt
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    ws.print_title_rows = "1:4"
    _wo_set_print_area(ws, "B", 4 + len(rows))


# ---------------------------------------------------------------------
# Throughput by Hour
# ---------------------------------------------------------------------

def _t_write_throughput_by_hour(wb, tickets):
    ws = wb.create_sheet("Throughput by Hour")
    note = (
        "Created and approved counts per hour-of-day. Throughput = "
        "jobs / unit time."
    )
    ws.cell(row=1, column=1, value="Throughput by Hour-of-Day").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Hour (UTC)", 10, "00"),
        ("Created", 10, FMT_INT),
        ("Approved", 10, FMT_INT),
        ("Per-minute (created)", 14, "0.000"),
        ("Per-30min (created)", 14, "0.00"),
        ("Per-hour (created)", 14, "0.00"),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    created_by_h = defaultdict(int)
    approved_by_h = defaultdict(int)
    for t in tickets:
        if t.created_at:
            created_by_h[t.created_at.hour] += 1
        a_at = getattr(t, "approved_at", None)
        if a_at:
            approved_by_h[a_at.hour] += 1

    for h in range(24):
        r = header_row + 1 + h
        c_count = created_by_h.get(h, 0)
        cells = [
            (1, h, "00"),
            (2, c_count, FMT_INT),
            (3, approved_by_h.get(h, 0), FMT_INT),
            (4, c_count / 60.0, "0.000"),
            (5, c_count / 2.0, "0.00"),
            (6, float(c_count), "0.00"),
        ]
        for col, val, fmt in cells:
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = header_row + 24
    ws.auto_filter.ref = f"A{header_row}:F{last_row}"
    ws.freeze_panes = ws[f"A{header_row + 1}"]
    _wo_data_bar(ws, "B", header_row + 1, last_row)
    _wo_data_bar(ws, "C", header_row + 1, last_row, color="64748B")
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "F", last_row)


# ---------------------------------------------------------------------
# Data Quality
# ---------------------------------------------------------------------

def _t_write_data_quality(wb, tickets):
    ws = wb.create_sheet("Data Quality")
    note = "Coverage and completeness of the ticket source data."
    ws.cell(row=1, column=1, value="Data Quality").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    n = len(tickets)
    miss_start = sum(1 for t in tickets if not getattr(t, "start_time", None))
    miss_end = sum(1 for t in tickets if not getattr(t, "end_time", None))
    miss_due = sum(1 for t in tickets if not getattr(t, "due_date", None))
    miss_vendor = sum(1 for t in tickets if not t.vendor_id)
    miss_contractor = sum(1 for t in tickets if not getattr(t, "assigned_contractor", None))

    seen = set()
    dup_ids = 0
    for t in tickets:
        if t.id is None:
            continue
        if t.id in seen:
            dup_ids += 1
        else:
            seen.add(t.id)

    rows = [
        ("Total ticket rows", n, FMT_INT),
        ("Missing Start Time", miss_start, FMT_INT),
        ("Missing End Time", miss_end, FMT_INT),
        ("Missing Due Date", miss_due, FMT_INT),
        ("Missing Vendor", miss_vendor, FMT_INT),
        ("Missing Assigned Contractor", miss_contractor, FMT_INT),
        ("Duplicate ticket ids", dup_ids, FMT_INT),
        ("% missing start_time", (miss_start / n) if n else 0, FMT_PCT),
        ("% missing end_time", (miss_end / n) if n else 0, FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.font = WO_KPI_VALUE_FONT
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.print_title_rows = "1:4"
    _wo_set_print_area(ws, "B", 4 + len(rows))


# =====================================================================
# Work Orders workbook — executive rebuild
# =====================================================================
#
# Tab order (Executive Summary first, set as active on save):
#   1. Executive Summary   — KPIs, top vendors, status mix, watch list
#   2. Work Orders Data    — flat source with derived duration columns
#   3. Vendor Scorecard    — per-vendor formulas vs. flat
#   4. Service Benchmarks  — per-service throughput + duration quartiles
#   5. Duration Analysis   — overall + per-service quartiles, outliers
#   6. Throughput by Hour  — created/completed counts by hour-of-day
#   7. Cost per Hour       — $/hr rates per vendor
#   8. Data Quality        — coverage / null counts / duplicates
#
# All scalar KPIs and per-vendor metrics are Excel formulas pointing at
# the flat tab, so editing rows in place recalculates the whole book.
# Distribution stats (quartiles, hourly counts) are precomputed in
# Python with a one-line note on each tab calling that out.

WO_FLAT_SHEET = "Work Orders Data"
WO_REPORT_TITLE = "Work Orders Report"
WO_HEADER_ROW = 4  # row carrying column titles on Work Orders Data

# Typography per spec
WO_TITLE_FONT = Font(name="Arial", size=18, bold=True, color="0F172A")
WO_TAB_TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F2937")
WO_HEADER_FONT_BOLD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
WO_BODY_FONT = Font(name="Arial", size=10, color="0F172A")
WO_KPI_VALUE_FONT = Font(name="Arial", size=16, bold=True, color="0F172A")
WO_KPI_LABEL_FONT = Font(name="Arial", size=9, color="6B7280")
WO_NOTE_FONT = Font(name="Arial", size=9, italic=True, color="6B7280")
WO_SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F3A8A")

# Muted, print-friendly fills
WO_HEADER_FILL = PatternFill("solid", fgColor="1F3A8A")          # navy
WO_SECTION_FILL = PatternFill("solid", fgColor="EEF2FF")         # ice
WO_BAND_FILL = PatternFill("solid", fgColor="F8FAFC")            # band
WO_GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")           # completed
WO_AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")           # in-progress
WO_RED_FILL = PatternFill("solid", fgColor="FEE2E2")             # halted/cancelled
WO_SLATE_FILL = PatternFill("solid", fgColor="F1F5F9")           # closed/neutral

# Number formats per spec
FMT_CURRENCY = '$#,##0;($#,##0);-'
FMT_CURRENCY_DEC = '$#,##0.00;($#,##0.00);-'
FMT_PCT = '0.0%;(0.0%);-'
FMT_HOURS = '#,##0.0'
FMT_INT = '#,##0;(#,##0);-'
FMT_DATE = 'yyyy-mm-dd'
FMT_DATETIME = 'yyyy-mm-dd hh:mm'

# Status colour map (red/amber/green/slate; grayscale-safe)
WO_STATUS_FILLS = {
    "COMPLETED": WO_GREEN_FILL,
    "CLOSED": WO_GREEN_FILL,
    "ASSIGNED": WO_AMBER_FILL,
    "IN_PROGRESS": WO_AMBER_FILL,
    "PENDING": WO_AMBER_FILL,
    "UNASSIGNED": WO_SLATE_FILL,
    "HALTED": WO_RED_FILL,
    "CANCELLED": WO_RED_FILL,
    "REJECTED": WO_RED_FILL,
}


def _wo_quoted(name):
    """Return sheet name quoted for use in formulas."""
    return f"'{name}'"


def _wo_print_setup(ws, period, tab_name, gridlines=True, report_title=None):
    """Apply the workbook-wide print spec to a single tab."""
    # Letter size + landscape if wide tab; orientation set by caller via
    # ws.page_setup.orientation before calling. Default landscape.
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    if not ws.page_setup.orientation:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.print_options.gridLines = gridlines
    ws.print_options.gridLinesSet = gridlines
    ws.print_options.horizontalCentered = True
    ws.oddHeader.left.text = report_title or WO_REPORT_TITLE
    ws.oddHeader.left.size = 9
    ws.oddHeader.center.text = tab_name
    ws.oddHeader.center.size = 9
    ws.oddHeader.right.text = f"Period: {period}"
    ws.oddHeader.right.size = 9
    ws.oddFooter.left.text = (
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    )
    ws.oddFooter.left.size = 9
    ws.oddFooter.center.text = "Confidential - Internal Use Only"
    ws.oddFooter.center.size = 9
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 9


def _wo_set_print_area(ws, last_col_letter, last_row):
    """Constrain the print area to actual data extents."""
    ws.print_area = f"A1:{last_col_letter}{last_row}"


def _wo_band_rows(ws, first_data_row, last_data_row, last_col):
    """Apply alternating row banding without using fills that override
    conditional formatting on status columns."""
    for r in range(first_data_row, last_data_row + 1):
        if (r - first_data_row) % 2 == 1:
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.value in (None, "00000000"):
                    cell.fill = WO_BAND_FILL


def _wo_apply_status_cf(ws, status_col_letter, first_row, last_row):
    """Red/amber/green status fills that print legibly in grayscale."""
    rng = f"{status_col_letter}{first_row}:{status_col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"COMPLETED"'], fill=WO_GREEN_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"CLOSED"'], fill=WO_GREEN_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"ASSIGNED"'], fill=WO_AMBER_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"IN_PROGRESS"'], fill=WO_AMBER_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"HALTED"'], fill=WO_RED_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"CANCELLED"'], fill=WO_RED_FILL),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"REJECTED"'], fill=WO_RED_FILL),
    )


def _wo_data_bar(ws, col_letter, first_row, last_row, color="64748B"):
    """Print-friendly slate data bar."""
    ws.conditional_formatting.add(
        f"{col_letter}{first_row}:{col_letter}{last_row}",
        DataBarRule(
            start_type="min",
            end_type="max",
            color=color,
            showValue=True,
            minLength=0,
            maxLength=100,
        ),
    )


def build_workorders_workbook(start=None, end=None):
    """Executive-quality work orders workbook.

    See the section banner above for tab order and design intent.
    """
    wb = Workbook()
    wb.remove(wb.active)

    work_orders = db.session.execute(select(WorkOrder)).scalars().all()
    if start or end:
        work_orders = [w for w in work_orders if _in_range(w.created_at, start, end)]
    vendors_by_id = {v.id: v for v in db.session.execute(select(Vendor)).scalars().all()}
    clients_by_id = {c.id: c for c in db.session.execute(select(Client)).scalars().all()}

    period = _period_label(start, end)
    n = len(work_orders)

    # Build flat data first so the other tabs can reference it.
    flat_last_data_row = _wo_write_flat(wb, work_orders, vendors_by_id, clients_by_id)

    _wo_write_exec_summary(wb, work_orders, vendors_by_id, period, flat_last_data_row)
    _wo_write_vendor_scorecard(wb, work_orders, vendors_by_id, flat_last_data_row)
    _wo_write_service_benchmarks(wb, work_orders)
    _wo_write_duration_analysis(wb, work_orders)
    _wo_write_throughput_by_hour(wb, work_orders)
    _wo_write_cost_per_hour(wb, work_orders, vendors_by_id, flat_last_data_row)
    _wo_write_data_quality(wb, work_orders, flat_last_data_row)

    # Move Executive Summary to position 0 (it's currently created after flat).
    exec_idx = wb.sheetnames.index("Executive Summary")
    if exec_idx != 0:
        wb.move_sheet("Executive Summary", offset=-exec_idx)

    wb.active = 0
    exec_ws = wb["Executive Summary"]
    exec_ws.sheet_view.zoomScale = 100
    exec_ws.sheet_view.selection[0].activeCell = "A1"
    exec_ws.sheet_view.selection[0].sqref = "A1"

    # Apply print setup last so every tab is configured.
    for ws in wb.worksheets:
        if ws.title == "Executive Summary":
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            _wo_print_setup(ws, period, ws.title, gridlines=False)
        else:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            _wo_print_setup(ws, period, ws.title, gridlines=True)

    return _to_workbook_bytes(wb)


# ---------------------------------------------------------------------
# 2. Work Orders Data (flat) — source of truth
# ---------------------------------------------------------------------

def _wo_write_flat(wb, work_orders, vendors_by_id, clients_by_id):
    ws = wb.create_sheet(WO_FLAT_SHEET)
    note = (
        "Source-of-truth flat list. Columns AC-AH are derived in-cell from the "
        "timestamps to the left (durations in hours, variance, lead time, on-time)."
    )
    ws.cell(row=1, column=1, value="Work Orders - source data").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3,
        column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Client", 26, None),
        ("Vendor", 28, None),
        ("WO #", 10, FMT_INT),
        ("Description", 38, None),
        ("Status", 14, None),
        ("Priority", 10, None),
        ("Service", 18, None),
        ("Location Type", 14, None),
        ("Location", 26, None),
        ("Latitude", 12, "#,##0.000000"),
        ("Longitude", 12, "#,##0.000000"),
        ("Well ID", 36, None),
        ("Est Cost", 14, FMT_CURRENCY),
        ("Est Quantity", 12, "#,##0.00"),
        ("Units", 8, None),
        ("Recurring?", 10, None),
        ("Recurrence", 12, None),
        ("Estimated Start", 16, FMT_DATE),
        ("Estimated End", 16, FMT_DATE),
        ("Assigned At", 16, FMT_DATETIME),
        ("Completed At", 16, FMT_DATETIME),
        ("Closed At", 16, FMT_DATETIME),
        ("Halted At", 16, FMT_DATETIME),
        ("Rejected At", 16, FMT_DATETIME),
        ("Created", 16, FMT_DATETIME),
        ("Updated", 16, FMT_DATETIME),
        ("Cancelled At", 16, FMT_DATETIME),
        ("Cancellation Reason", 28, None),
        ("Actual Duration (hrs)", 14, FMT_HOURS),
        ("Estimated Duration (hrs)", 14, FMT_HOURS),
        ("Variance (hrs)", 12, FMT_HOURS),
        ("Variance %", 12, FMT_PCT),
        ("Lead Time (hrs)", 14, FMT_HOURS),
        ("On Time?", 10, None),
    ]
    _apply_widths(ws, cols)

    header_row = WO_HEADER_ROW
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    sorted_wos = sorted(
        work_orders,
        key=lambda w: (
            _vendor_name(vendors_by_id.get(w.assigned_vendor)),
            -(w.created_at.timestamp() if w.created_at else 0),
        ),
    )
    for offset, w in enumerate(sorted_wos):
        row = header_row + 1 + offset
        v = vendors_by_id.get(w.assigned_vendor)
        c = clients_by_id.get(w.client_id) if w.client_id else None
        client_name = getattr(c, "client_name", None) if c else None
        service = w.service.service if getattr(w, "service", None) else None
        location_str = getattr(w, "location", None) or (
            f"{w.latitude}, {w.longitude}"
            if getattr(w, "latitude", None) is not None and getattr(w, "longitude", None) is not None
            else ""
        )
        lat = getattr(w, "latitude", None)
        lng = getattr(w, "longitude", None)

        # T = Assigned At (col 20), U = Completed At (col 21), R = Estimated Start (col 18),
        # S = Estimated End (col 19), Y = Created (col 25)
        actual_dur = (
            f'=IF(AND(T{row}<>"",U{row}<>""),(U{row}-T{row})*24,"")'
        )
        est_dur = (
            f'=IF(AND(R{row}<>"",S{row}<>""),(S{row}-R{row})*24,"")'
        )
        variance = (
            f'=IF(AND(AC{row}<>"",AD{row}<>""),AC{row}-AD{row},"")'
        )
        variance_pct = (
            f'=IF(AND(AC{row}<>"",AD{row}<>"",AD{row}>0),'
            f'(AC{row}-AD{row})/AD{row},"")'
        )
        lead_time = (
            f'=IF(AND(Y{row}<>"",T{row}<>""),(T{row}-Y{row})*24,"")'
        )
        on_time = (
            f'=IF(AND(U{row}<>"",S{row}<>""),'
            f'IF(U{row}<=S{row},"Yes","No"),"")'
        )

        values = [
            client_name or "",
            _vendor_name(v),
            getattr(w, "work_order_code", None),
            w.description or "",
            _enum_value(getattr(w, "current_status", None)),
            _enum_value(getattr(w, "priority", None)),
            service or "",
            _enum_value(getattr(w, "location_type", None)) or "",
            location_str,
            float(lat) if lat is not None else None,
            float(lng) if lng is not None else None,
            getattr(w, "well_id", None) or "",
            float(getattr(w, "estimated_cost", 0) or 0),
            float(getattr(w, "estimated_quantity", 0) or 0),
            getattr(w, "units", None) or "",
            "Yes" if getattr(w, "is_recurring", False) else "No",
            _enum_value(getattr(w, "recurrence_type", None)) or "",
            _naive(getattr(w, "estimated_start_date", None)),
            _naive(getattr(w, "estimated_end_date", None)),
            _naive(getattr(w, "assigned_at", None)),
            _naive(getattr(w, "completed_at", None)),
            _naive(getattr(w, "closed_at", None)),
            _naive(getattr(w, "halted_at", None)),
            _naive(getattr(w, "rejected_at", None)),
            _naive(w.created_at),
            _naive(getattr(w, "updated_at", None)),
            _naive(getattr(w, "cancelled_at", None)),
            getattr(w, "cancellation_reason", None) or "",
            actual_dur,
            est_dur,
            variance,
            variance_pct,
            lead_time,
            on_time,
        ]

        for c_idx, ((_, _w, fmt), value) in enumerate(zip(cols, values), start=1):
            cell = ws.cell(row=row, column=c_idx, value=value)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_data_row = header_row + len(sorted_wos)
    last_col = len(cols)
    last_col_letter = get_column_letter(last_col)

    if sorted_wos:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_apply_status_cf(ws, "E", header_row + 1, last_data_row)
        _wo_data_bar(ws, "M", header_row + 1, last_data_row)
        _wo_data_bar(ws, "AC", header_row + 1, last_data_row, color="334155")
        # Highlight rows where the variance ratio is >= 50% over estimate.
        ws.conditional_formatting.add(
            f"AF{header_row + 1}:AF{last_data_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=WO_RED_FILL),
        )

    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, last_col_letter, max(last_data_row, header_row))
    return last_data_row


# ---------------------------------------------------------------------
# 1. Executive Summary
# ---------------------------------------------------------------------

def _wo_write_exec_summary(wb, work_orders, vendors_by_id, period, flat_last_row):
    ws = wb.create_sheet("Executive Summary")
    flat = _wo_quoted(WO_FLAT_SHEET)
    n = len(work_orders)

    # Layout:
    #   Row 1: report title (Arial 18 bold)
    #   Row 2: period
    #   Row 3: generated timestamp + record count
    #   Row 4: blank
    #   Row 5: "what changed" paragraph (omitted cleanly if no prior period)
    #   Row 7+: KPI cards (label/value pairs in a grid)
    #   Then Status mix, Top 5 vendors by spend, Watch list

    ws.merge_cells("A1:F1")
    ws["A1"] = WO_REPORT_TITLE
    ws["A1"].font = WO_TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = f"Period: {period}    ·    {n:,} work order(s)"
    ws["A2"].font = WO_NOTE_FONT

    ws["A3"] = (
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A3"].font = WO_NOTE_FONT

    # KPI grid (no merged cells inside data ranges; merges are in title only)
    kpi_row = 6
    kpi_specs = [
        ("Total work orders",
         f"=COUNTA({flat}!C{WO_HEADER_ROW + 1}:C{flat_last_row})", FMT_INT),
        ("Total estimated cost",
         f"=SUM({flat}!M{WO_HEADER_ROW + 1}:M{flat_last_row})", FMT_CURRENCY),
        ("Average actual duration (hrs)",
         f'=IFERROR(AVERAGE({flat}!AC{WO_HEADER_ROW + 1}:AC{flat_last_row}),0)',
         FMT_HOURS),
        ("Average variance vs estimate",
         f'=IFERROR(AVERAGE({flat}!AF{WO_HEADER_ROW + 1}:AF{flat_last_row}),0)',
         FMT_PCT),
        ("On-time completion rate",
         f'=IFERROR(COUNTIF({flat}!AH{WO_HEADER_ROW + 1}:AH{flat_last_row},"Yes")'
         f'/(COUNTIF({flat}!AH{WO_HEADER_ROW + 1}:AH{flat_last_row},"Yes")'
         f'+COUNTIF({flat}!AH{WO_HEADER_ROW + 1}:AH{flat_last_row},"No")),0)',
         FMT_PCT),
        ("Halted / cancelled count",
         f'=COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"HALTED")'
         f'+COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"CANCELLED")',
         FMT_INT),
    ]

    # 3 columns x 2 rows of KPI cards
    for idx, (label, formula, fmt) in enumerate(kpi_specs):
        col = (idx % 3) * 2 + 1   # cols 1, 3, 5
        row = kpi_row + (idx // 3) * 3
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.font = WO_KPI_LABEL_FONT
        value_cell = ws.cell(row=row + 1, column=col, value=formula)
        value_cell.font = WO_KPI_VALUE_FONT
        value_cell.number_format = fmt
        for c in (label_cell, value_cell):
            c.border = THIN_BORDER

    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 22

    # ----- Status breakdown -----
    section_row = kpi_row + 7
    ws.cell(row=section_row, column=1, value="STATUS BREAKDOWN").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    cols = [("Status", 22, None), ("Count", 12, FMT_INT), ("Share", 10, FMT_PCT)]
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD
        h.fill = WO_HEADER_FILL

    statuses = [
        "UNASSIGNED", "PENDING", "ASSIGNED", "IN_PROGRESS", "COMPLETED",
        "CLOSED", "HALTED", "CANCELLED", "REJECTED",
    ]
    total_formula = f"COUNTA({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row})"
    for i, status in enumerate(statuses):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=status).font = WO_BODY_FONT
        ws.cell(
            row=r, column=2,
            value=f'=COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"{status}")',
        ).number_format = FMT_INT
        ws.cell(
            row=r, column=3,
            value=f'=IFERROR(COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"{status}")/{total_formula},0)',
        ).number_format = FMT_PCT
    status_last = section_row + len(statuses)

    # ----- Top 5 vendors by spend (precomputed; vendor names live on flat) -----
    section_row = status_last + 3
    ws.cell(row=section_row, column=1, value="TOP 5 VENDORS BY ESTIMATED SPEND").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    for c_idx, label in enumerate(["Vendor", "WO Count", "Est Total"], start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD
        h.fill = WO_HEADER_FILL

    spend_by_vendor = defaultdict(lambda: {"count": 0, "value": 0.0, "name": ""})
    for w in work_orders:
        if not w.assigned_vendor:
            continue
        slot = spend_by_vendor[w.assigned_vendor]
        slot["count"] += 1
        slot["value"] += float(getattr(w, "estimated_cost", 0) or 0)
        slot["name"] = _vendor_name(vendors_by_id.get(w.assigned_vendor))
    top_vendors = sorted(spend_by_vendor.values(), key=lambda x: x["value"], reverse=True)[:5]
    for i, slot in enumerate(top_vendors):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=slot["name"]).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=slot["count"]).number_format = FMT_INT
        ws.cell(row=r, column=3, value=slot["value"]).number_format = FMT_CURRENCY
    top_last = section_row + max(len(top_vendors), 1)

    # ----- Watch list: halted / cancelled / >7d open / variance >= 50% -----
    section_row = top_last + 3
    ws.cell(row=section_row, column=1, value="WATCH LIST").font = WO_SECTION_FONT
    ws.cell(row=section_row, column=1).fill = WO_SECTION_FILL
    section_row += 1
    watch_cols = [("Indicator", 38, None), ("Count", 12, FMT_INT)]
    for c_idx, (label, _w, _f) in enumerate(watch_cols, start=1):
        h = ws.cell(row=section_row, column=c_idx, value=label)
        h.font = WO_HEADER_FONT_BOLD
        h.fill = WO_HEADER_FILL

    watch_specs = [
        ("Work orders currently HALTED",
         f'=COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"HALTED")'),
        ("Work orders currently CANCELLED",
         f'=COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"CANCELLED")'),
        ("Work orders REJECTED",
         f'=COUNTIF({flat}!E{WO_HEADER_ROW + 1}:E{flat_last_row},"REJECTED")'),
        ("Late completions (On Time = No)",
         f'=COUNTIF({flat}!AH{WO_HEADER_ROW + 1}:AH{flat_last_row},"No")'),
        ("Variance >= 50% over estimate",
         f'=COUNTIFS({flat}!AF{WO_HEADER_ROW + 1}:AF{flat_last_row},">=0.5")'),
    ]
    for i, (label, formula) in enumerate(watch_specs):
        r = section_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        ws.cell(row=r, column=2, value=formula).number_format = FMT_INT


# ---------------------------------------------------------------------
# 3. Vendor Scorecard
# ---------------------------------------------------------------------

def _wo_write_vendor_scorecard(wb, work_orders, vendors_by_id, flat_last_row):
    ws = wb.create_sheet("Vendor Scorecard")
    flat = _wo_quoted(WO_FLAT_SHEET)
    note = (
        "Per-vendor performance vs the flat data tab. All metrics are formulas; "
        "edit rows in 'Work Orders Data' and this tab recalculates."
    )
    ws.cell(row=1, column=1, value="Vendor Scorecard").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Vendor", 30, None),
        ("WO Count", 10, FMT_INT),
        ("Total Est Cost", 14, FMT_CURRENCY),
        ("Avg Actual Duration (hrs)", 16, FMT_HOURS),
        ("Avg Estimated Duration (hrs)", 16, FMT_HOURS),
        ("Avg Variance %", 14, FMT_PCT),
        ("On-Time %", 12, FMT_PCT),
        ("Halt/Cancel Count", 14, FMT_INT),
        ("Halt/Cancel Rate", 14, FMT_PCT),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    vendor_names = sorted(
        {_vendor_name(vendors_by_id.get(w.assigned_vendor)) for w in work_orders if w.assigned_vendor}
    )

    flat_rows = (WO_HEADER_ROW + 1, flat_last_row)
    a, b = flat_rows
    for i, name in enumerate(vendor_names):
        r = header_row + 1 + i
        v_ref = f'A{r}'
        ws.cell(row=r, column=1, value=name).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({flat}!B{a}:B{b},{v_ref})').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIF({flat}!B{a}:B{b},{v_ref},{flat}!M{a}:M{b})').number_format = FMT_CURRENCY
        ws.cell(row=r, column=4,
                value=f'=IFERROR(AVERAGEIFS({flat}!AC{a}:AC{b},{flat}!B{a}:B{b},{v_ref},{flat}!AC{a}:AC{b},">0"),0)'
                ).number_format = FMT_HOURS
        ws.cell(row=r, column=5,
                value=f'=IFERROR(AVERAGEIFS({flat}!AD{a}:AD{b},{flat}!B{a}:B{b},{v_ref},{flat}!AD{a}:AD{b},">0"),0)'
                ).number_format = FMT_HOURS
        ws.cell(row=r, column=6,
                value=f'=IFERROR(AVERAGEIFS({flat}!AF{a}:AF{b},{flat}!B{a}:B{b},{v_ref},{flat}!AF{a}:AF{b},"<>"),0)'
                ).number_format = FMT_PCT
        ws.cell(row=r, column=7,
                value=(
                    f'=IFERROR('
                    f'COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!AH{a}:AH{b},"Yes")'
                    f'/(COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!AH{a}:AH{b},"Yes")'
                    f'+COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!AH{a}:AH{b},"No")),0)'
                )).number_format = FMT_PCT
        ws.cell(row=r, column=8,
                value=(
                    f'=COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"HALTED")'
                    f'+COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"CANCELLED")'
                )).number_format = FMT_INT
        ws.cell(row=r, column=9,
                value=(
                    f'=IFERROR(('
                    f'COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"HALTED")'
                    f'+COUNTIFS({flat}!B{a}:B{b},{v_ref},{flat}!E{a}:E{b},"CANCELLED"))'
                    f'/COUNTIF({flat}!B{a}:B{b},{v_ref}),0)'
                )).number_format = FMT_PCT
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    last_row = header_row + len(vendor_names)
    if vendor_names:
        ws.auto_filter.ref = f"A{header_row}:I{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        # Halt/cancel rate red bar
        ws.conditional_formatting.add(
            f"I{header_row + 1}:I{last_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.1"], fill=WO_RED_FILL),
        )
        _wo_data_bar(ws, "C", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "I", max(last_row, header_row))


# ---------------------------------------------------------------------
# 4. Service Type Benchmarks
# ---------------------------------------------------------------------

def _wo_durations_for(filter_fn, work_orders):
    out = []
    for w in work_orders:
        if not filter_fn(w):
            continue
        a = getattr(w, "assigned_at", None)
        c = getattr(w, "completed_at", None)
        if a and c:
            hrs = (c - a).total_seconds() / 3600.0
            if hrs > 0:
                out.append(hrs)
    return out


def _quartiles(values):
    if not values:
        return (0, 0, 0, 0, 0, 0)
    if len(values) == 1:
        v = values[0]
        return (v, v, v, v, v, v)
    sorted_vals = sorted(values)
    try:
        q = quantiles(sorted_vals, n=4, method="inclusive")
        p25, med, p75 = q[0], q[1], q[2]
    except Exception:
        med = median(sorted_vals)
        p25 = sorted_vals[len(sorted_vals) // 4]
        p75 = sorted_vals[(3 * len(sorted_vals)) // 4]
    # p90
    try:
        p90 = quantiles(sorted_vals, n=10, method="inclusive")[8]
    except Exception:
        p90 = sorted_vals[int(len(sorted_vals) * 0.9)]
    return (sorted_vals[0], p25, med, p75, p90, sorted_vals[-1])


def _wo_write_service_benchmarks(wb, work_orders):
    ws = wb.create_sheet("Service Benchmarks")
    note = (
        "Per service type: count, total estimated cost, throughput vs the "
        "period, and duration quartiles. Quartiles are computed in Python "
        "from the data shipped in this workbook (snapshot at generation)."
    )
    ws.cell(row=1, column=1, value="Service Type Benchmarks").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Service", 24, None),
        ("Count", 10, FMT_INT),
        ("Total Est Cost", 16, FMT_CURRENCY),
        ("Avg Duration (hrs)", 16, FMT_HOURS),
        ("Min", 8, FMT_HOURS),
        ("p25", 8, FMT_HOURS),
        ("Median", 10, FMT_HOURS),
        ("p75", 8, FMT_HOURS),
        ("p90", 8, FMT_HOURS),
        ("Max", 10, FMT_HOURS),
        ("Outliers (>2x median)", 16, FMT_INT),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    services = sorted(
        {(w.service.service if getattr(w, "service", None) else "(no service)")
         for w in work_orders}
    )
    row = header_row
    for svc in services:
        row += 1
        same = [w for w in work_orders
                if (getattr(w.service, "service", None) if getattr(w, "service", None) else "(no service)") == svc]
        durs = _wo_durations_for(
            lambda w, _svc=svc: ((getattr(w.service, "service", None) if getattr(w, "service", None) else "(no service)") == _svc),
            work_orders,
        )
        total_cost = sum(float(getattr(w, "estimated_cost", 0) or 0) for w in same)
        avg = sum(durs) / len(durs) if durs else 0
        mn, p25, med, p75, p90, mx = _quartiles(durs)
        outliers = sum(1 for d in durs if med and d > 2 * med)
        values = [svc, len(same), total_cost, avg, mn, p25, med, p75, p90, mx, outliers]
        for c_idx, ((_, _w, fmt), val) in enumerate(zip(cols, values), start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            if fmt:
                cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = row
    if services:
        ws.auto_filter.ref = f"A{header_row}:K{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "C", header_row + 1, last_row)
        _wo_data_bar(ws, "G", header_row + 1, last_row, color="334155")
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "K", max(last_row, header_row))


# ---------------------------------------------------------------------
# 5. Duration Analysis (overall)
# ---------------------------------------------------------------------

def _wo_write_duration_analysis(wb, work_orders):
    ws = wb.create_sheet("Duration Analysis")
    note = (
        "Overall duration distribution across all work orders with both "
        "timestamps. Outliers are jobs longer than 2x the overall median."
    )
    ws.cell(row=1, column=1, value="Duration Analysis").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    durs = _wo_durations_for(lambda w: True, work_orders)
    mn, p25, med, p75, p90, mx = _quartiles(durs)
    avg = sum(durs) / len(durs) if durs else 0
    excluded = len(work_orders) - len(durs)

    rows = [
        ("Jobs included (have assigned + completed)", len(durs), FMT_INT),
        ("Jobs excluded (open or never assigned)", excluded, FMT_INT),
        ("Average duration (hrs)", avg, FMT_HOURS),
        ("Min", mn, FMT_HOURS),
        ("p25", p25, FMT_HOURS),
        ("Median", med, FMT_HOURS),
        ("p75", p75, FMT_HOURS),
        ("p90", p90, FMT_HOURS),
        ("Max", mx, FMT_HOURS),
        ("Outliers (>2x median)", sum(1 for d in durs if med and d > 2 * med), FMT_INT),
    ]
    for c_idx, (label, val, fmt) in enumerate(rows):
        r = 5 + c_idx
        l = ws.cell(row=r, column=1, value=label)
        l.font = WO_BODY_FONT
        v = ws.cell(row=r, column=2, value=val)
        v.font = WO_KPI_VALUE_FONT
        v.number_format = fmt
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    ws.print_title_rows = "1:4"
    _wo_set_print_area(ws, "B", 4 + len(rows))


# ---------------------------------------------------------------------
# 6. Throughput by Hour
# ---------------------------------------------------------------------

def _wo_write_throughput_by_hour(wb, work_orders):
    ws = wb.create_sheet("Throughput by Hour")
    note = (
        "Created and completed counts per hour-of-day across the period, "
        "showing when activity is concentrated. Throughput = jobs / unit time."
    )
    ws.cell(row=1, column=1, value="Throughput by Hour-of-Day").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Hour (UTC)", 10, "00"),
        ("Created", 10, FMT_INT),
        ("Completed", 10, FMT_INT),
        ("Per-minute (created)", 14, "0.000"),
        ("Per-30min (created)", 14, "0.00"),
        ("Per-hour (created)", 14, "0.00"),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    created_by_h = defaultdict(int)
    completed_by_h = defaultdict(int)
    for w in work_orders:
        if w.created_at:
            created_by_h[w.created_at.hour] += 1
        completed = getattr(w, "completed_at", None)
        if completed:
            completed_by_h[completed.hour] += 1

    for h in range(24):
        r = header_row + 1 + h
        c_count = created_by_h.get(h, 0)
        cells = [
            (1, h, "00"),
            (2, c_count, FMT_INT),
            (3, completed_by_h.get(h, 0), FMT_INT),
            (4, c_count / 60.0, "0.000"),
            (5, c_count / 2.0, "0.00"),
            (6, float(c_count), "0.00"),
        ]
        for col, val, fmt in cells:
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            cell.font = WO_BODY_FONT
            cell.border = THIN_BORDER

    last_row = header_row + 24
    ws.auto_filter.ref = f"A{header_row}:F{last_row}"
    ws.freeze_panes = ws[f"A{header_row + 1}"]
    _wo_data_bar(ws, "B", header_row + 1, last_row)
    _wo_data_bar(ws, "C", header_row + 1, last_row, color="64748B")
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "F", last_row)


# ---------------------------------------------------------------------
# 7. Cost per Hour (per vendor)
# ---------------------------------------------------------------------

def _wo_write_cost_per_hour(wb, work_orders, vendors_by_id, flat_last_row):
    ws = wb.create_sheet("Cost per Hour")
    flat = _wo_quoted(WO_FLAT_SHEET)
    note = (
        "$/hr by vendor: total estimated cost / total actual duration hours "
        "(jobs without both assigned + completed timestamps are excluded). "
        "Per-30min and per-minute rates derived from $/hr."
    )
    ws.cell(row=1, column=1, value="Cost per Unit Time").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    cols = [
        ("Vendor", 30, None),
        ("Total Est Cost", 16, FMT_CURRENCY),
        ("Total Hours", 12, FMT_HOURS),
        ("$/hour", 12, FMT_CURRENCY_DEC),
        ("$/30min", 12, FMT_CURRENCY_DEC),
        ("$/minute", 12, FMT_CURRENCY_DEC),
    ]
    _apply_widths(ws, cols)
    header_row = 4
    for c_idx, (label, _w, _f) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=label)
        cell.font = WO_HEADER_FONT_BOLD
        cell.fill = WO_HEADER_FILL
        cell.border = THIN_BORDER

    vendor_names = sorted(
        {_vendor_name(vendors_by_id.get(w.assigned_vendor))
         for w in work_orders if w.assigned_vendor}
    )
    a, b = WO_HEADER_ROW + 1, flat_last_row
    for i, name in enumerate(vendor_names):
        r = header_row + 1 + i
        v_ref = f"A{r}"
        ws.cell(row=r, column=1, value=name).font = WO_BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=SUMIF({flat}!B{a}:B{b},{v_ref},{flat}!M{a}:M{b})').number_format = FMT_CURRENCY
        ws.cell(row=r, column=3,
                value=f'=IFERROR(SUMIFS({flat}!AC{a}:AC{b},{flat}!B{a}:B{b},{v_ref},{flat}!AC{a}:AC{b},">0"),0)'
                ).number_format = FMT_HOURS
        ws.cell(row=r, column=4,
                value=f'=IFERROR(B{r}/C{r},0)').number_format = FMT_CURRENCY_DEC
        ws.cell(row=r, column=5,
                value=f'=D{r}/2').number_format = FMT_CURRENCY_DEC
        ws.cell(row=r, column=6,
                value=f'=D{r}/60').number_format = FMT_CURRENCY_DEC
        for c_idx in range(1, len(cols) + 1):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    last_row = header_row + len(vendor_names)
    if vendor_names:
        ws.auto_filter.ref = f"A{header_row}:F{last_row}"
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        _wo_data_bar(ws, "D", header_row + 1, last_row)
    ws.print_title_rows = f"1:{header_row}"
    _wo_set_print_area(ws, "F", max(last_row, header_row))


# ---------------------------------------------------------------------
# 8. Data Quality
# ---------------------------------------------------------------------

def _wo_write_data_quality(wb, work_orders, flat_last_row):
    ws = wb.create_sheet("Data Quality")
    note = (
        "Coverage and completeness of the source data driving the report."
    )
    ws.cell(row=1, column=1, value="Data Quality").font = WO_TAB_TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = WO_NOTE_FONT
    ws.cell(
        row=3, column=1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    ).font = WO_NOTE_FONT

    n = len(work_orders)
    miss_assigned = sum(1 for w in work_orders if not getattr(w, "assigned_at", None))
    miss_completed = sum(1 for w in work_orders if not getattr(w, "completed_at", None))
    miss_estimated_dates = sum(
        1 for w in work_orders
        if not getattr(w, "estimated_start_date", None) or not getattr(w, "estimated_end_date", None)
    )
    miss_cost = sum(1 for w in work_orders if not getattr(w, "estimated_cost", None))
    miss_vendor = sum(1 for w in work_orders if not w.assigned_vendor)

    seen = set()
    dup_codes = 0
    for w in work_orders:
        code = getattr(w, "work_order_code", None)
        if code is None:
            continue
        if code in seen:
            dup_codes += 1
        else:
            seen.add(code)

    rows = [
        ("Total work order rows", n, FMT_INT),
        ("Missing Assigned At", miss_assigned, FMT_INT),
        ("Missing Completed At", miss_completed, FMT_INT),
        ("Missing Estimated Start/End", miss_estimated_dates, FMT_INT),
        ("Missing Estimated Cost", miss_cost, FMT_INT),
        ("Missing Assigned Vendor", miss_vendor, FMT_INT),
        ("Duplicate work_order_code values", dup_codes, FMT_INT),
        ("% missing assigned timestamps", (miss_assigned / n) if n else 0, FMT_PCT),
        ("% missing completed timestamps", (miss_completed / n) if n else 0, FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = WO_BODY_FONT
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.font = WO_KPI_VALUE_FONT
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.print_title_rows = "1:4"
    _wo_set_print_area(ws, "B", 4 + len(rows))



# ---------------------------------------------------------------------------
# Vendors workbook
# ---------------------------------------------------------------------------


def build_vendors_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    vendors = db.session.execute(
        select(Vendor).options(joinedload(Vendor.address))
    ).unique().scalars().all()

    _write_vendors_summary_sheet(wb, vendors)
    _write_vendors_directory_sheet(wb, vendors)
    _write_vendor_services_sheet(wb, vendors)
    return _to_workbook_bytes(wb)


def _write_vendors_summary_sheet(ws_wb, vendors):
    ws = ws_wb.create_sheet("Summary")
    _write_title(ws, "Vendor Directory Summary", f"{len(vendors)} vendor(s) on file.")

    by_status = defaultdict(int)
    by_compliance = defaultdict(int)
    onboarded = 0
    for v in vendors:
        by_status[_enum_value(getattr(v, "status", None)) or "UNKNOWN"] += 1
        by_compliance[_enum_value(getattr(v, "compliance_status", None)) or "UNKNOWN"] += 1
        if getattr(v, "onboarding", False):
            onboarded += 1

    row = 5
    ws.cell(row=row, column=1, value="HEADCOUNT").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    row += 1
    cols = [("Metric", 28, None), ("Count", 12, "#,##0")]
    _apply_widths(ws, cols)
    _write_header(ws, row, cols)
    for label, value in [
        ("Total vendors", len(vendors)),
        ("Onboarded vendors", onboarded),
    ]:
        row += 1
        _write_row(ws, row, cols, [label, value])

    row += 3
    ws.cell(row=row, column=1, value="BY STATUS").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    row += 1
    _write_header(ws, row, cols)
    for status, count in sorted(by_status.items(), key=lambda kv: kv[1], reverse=True):
        row += 1
        _write_row(ws, row, cols, [status, count], status_col=0)

    row += 3
    ws.cell(row=row, column=1, value="BY COMPLIANCE").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    row += 1
    _write_header(ws, row, cols)
    for compliance, count in sorted(by_compliance.items(), key=lambda kv: kv[1], reverse=True):
        row += 1
        _write_row(ws, row, cols, [compliance, count])


def _write_vendors_directory_sheet(ws_wb, vendors):
    ws = ws_wb.create_sheet("Vendors")
    _write_title(ws, "Vendor Directory", "All vendors with contact, status, and compliance details.")

    header_row = 5
    cols = [
        ("Company Name", 30, None),
        ("Vendor Code", 14, None),
        ("Status", 12, None),
        ("Compliance", 14, None),
        ("Onboarding", 12, None),
        ("Primary Contact", 24, None),
        ("Email", 28, None),
        ("Phone", 16, None),
        ("Address", 36, None),
        ("Description", 40, None),
        ("Start Date", 14, "yyyy-mm-dd"),
        ("End Date", 14, "yyyy-mm-dd"),
        ("Created", 18, "yyyy-mm-dd hh:mm"),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)

    sorted_vendors = sorted(vendors, key=_vendor_name)
    for i, v in enumerate(sorted_vendors, start=header_row + 1):
        addr_str = ""
        if v.address:
            parts = [
                v.address.street,
                v.address.city,
                v.address.state,
                v.address.zip,
            ]
            addr_str = ", ".join(p for p in parts if p)
        _write_row(
            ws,
            i,
            cols,
            [
                _vendor_name(v),
                v.vendor_code or "",
                _enum_value(getattr(v, "status", None)),
                _enum_value(getattr(v, "compliance_status", None)),
                "Yes" if getattr(v, "onboarding", False) else "No",
                v.primary_contact_name or "",
                v.company_email or "",
                v.company_phone or "",
                addr_str,
                v.description or "",
                _naive(getattr(v, "start_date", None)),
                _naive(getattr(v, "end_date", None)),
                _naive(v.created_at),
            ],
            status_col=2,
        )
    _finalize_data_sheet(ws, header_row, cols, len(sorted_vendors))


def _write_vendor_services_sheet(ws_wb, vendors):
    ws = ws_wb.create_sheet("Vendor Services")
    _write_title(ws, "Vendor Services", "One row per vendor-service link.")

    header_row = 5
    cols = [
        ("Vendor", 30, None),
        ("Service", 26, None),
    ]
    _apply_widths(ws, cols)
    _write_header(ws, header_row, cols)

    rows = []
    for v in vendors:
        services = getattr(v, "vendor_services", None) or []
        for vs in services:
            svc = getattr(vs, "service", None) or getattr(vs, "service_type", None)
            svc_name = getattr(svc, "service", None) if svc else None
            if svc_name:
                rows.append([_vendor_name(v), svc_name])
    rows.sort(key=lambda r: (r[0], r[1]))
    for i, row in enumerate(rows, start=header_row + 1):
        _write_row(ws, i, cols, row)
    _finalize_data_sheet(ws, header_row, cols, len(rows))


# ---------------------------------------------------------------------------
# Misc
# ---------------------------------------------------------------------------


def parse_date_range(from_str, to_str):
    """Parse 'YYYY-MM-DD' strings into (datetime, datetime) tuple. Returns
    (None, None) when both blank. The 'to' date becomes end-of-day so the
    full day is included.
    """
    start = None
    end = None
    if from_str:
        try:
            start = datetime.strptime(from_str, "%Y-%m-%d")
        except ValueError:
            start = None
    if to_str:
        try:
            d = datetime.strptime(to_str, "%Y-%m-%d")
            end = d.replace(hour=23, minute=59, second=59)
        except ValueError:
            end = None
    return start, end


def _in_range(value, start, end):
    if value is None:
        return False
    v = _naive(value)
    if start and v < start:
        return False
    if end and v > end:
        return False
    return True


def _period_label(start, end):
    if not start and not end:
        return "All time"
    if start and end:
        return f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"
    if start:
        return f"From {start.strftime('%Y-%m-%d')}"
    return f"Through {end.strftime('%Y-%m-%d')}"


def _enum_value(value):
    if value is None:
        return None
    if hasattr(value, "value"):
        v = value.value
        return v.upper() if isinstance(v, str) else v
    return str(value)
